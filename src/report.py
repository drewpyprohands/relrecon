"""
Report generation module for the relational matching framework.

Generates formatted Excel workbooks from matching pipeline results:
- Summary tab: recipe config, per-step counts, cascade explanation
- Matched tab: matched records with derived fields, confidence, validation columns
- Analysis tab: unmatched records with reason codes

Uses openpyxl for Excel formatting (headers, conditional formatting, column widths).
"""

from pathlib import Path
from typing import Optional

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Conditional fill colors for address scores
SCORE_HIGH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
SCORE_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Yellow
SCORE_LOW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Red

ANALYSIS_HEADER_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

# compare_columns observations. Deliberately inverted against the score
# banding above: for an id comparison, "higher" is the worse outcome.
CMP_HIGHER = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
CMP_LOWER = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Light green


# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# Main Report columns (order matters for Excel output)
MAIN_REPORT_COLUMNS = [
    ("l3_fmly_nm", "Source L3 Name"),
    ("vendor_id", "Source Vendor ID"),
    ("tpty_assm_nm", "Assessment Name"),
    ("hq_addr1", "Source Address 1"),
    ("hq_addr2", "Source Address 2"),
    ("derived_l1_name", "Derived L1 Name"),
    ("derived_l1_id", "Derived L1 ID"),
    ("match_step", "Match Source"),
    ("match_tier", "Match Tier"),
    ("name_score", "Name Score"),
    ("addr_score", "Address Score"),
    ("addr_street_match", "Street Match"),
    ("addr_comparison", "Address Comparison"),
    ("addr_tier", "Address Tier"),
]

# Destination fields (check for suffixed columns from join)
# Multiple entries per header handle core_parent vs Pop3 column naming
DEST_COLUMNS = [
    ("Vendor Name", "Dest L3 Name"),
    ("Vendor Name_dst", "Dest L3 Name"),
    ("l3_fmly_nm_dst", "Dest L3 Name"),
    ("Address1", "Dest Address 1"),
    ("hq_addr1_dst", "Dest Address 1"),
    ("Address2", "Dest Address 2"),
    ("hq_addr2_dst", "Dest Address 2"),
]

ANALYSIS_COLUMNS = [
    ("l3_fmly_nm", "L3 Name"),
    ("vendor_id", "Vendor ID"),
    ("tpty_assm_nm", "Assessment Name"),
    ("hq_addr1", "Address 1"),
    ("hq_addr2", "Address 2"),
    ("l1_fmly_nm", "L1 Name (invalid)"),
    ("tpty_l1_id", "L1 ID (invalid)"),
    ("data_entry_type", "Data Entry Type"),
    ("rq_intk_user", "Request User"),
    ("reason_code", "Reason Code"),
    ("rejection_step", "Rejection Step"),
    ("best_rejected_score", "Best Rejected Score"),
]


# ---------------------------------------------------------------------------
# Sheet writing helpers
# ---------------------------------------------------------------------------

def _write_headers(ws, columns: list, header_fill=HEADER_FILL):
    """Write formatted headers to a worksheet."""
    for col_idx, (_, header) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _write_data(ws, df: pl.DataFrame, columns: list, start_row: int = 2,
                compare_cols: set | None = None):
    """Write DataFrame rows to worksheet, matching columns by name.

    Uses iter_rows. Required for cell-by-cell openpyxl writes.
    Not a data processing loop (ADR-001 prohibits iterrows for data ops,
    not for output serialization).

    ``compare_cols`` names the compare_columns observation columns (by field
    and by header, since the two write paths differ) so they get their own
    higher/lower banding.
    """
    available_cols = set(df.columns)

    for row_idx, row in enumerate(df.iter_rows(named=True), start_row):
        for col_idx, (col_name, _) in enumerate(columns, 1):
            if col_name in available_cols:
                value = row[col_name]
                # Round score columns for clean display
                if col_name in ("addr_score", "name_score", "addr_street_score",
                                "best_rejected_score") and value is not None:
                    try:
                        value = round(float(value), 2)
                    except (ValueError, TypeError):
                        pass
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER

                # Conditional formatting for score columns (0-100 scale)
                if col_name in ("addr_score", "name_score") and value is not None:
                    try:
                        score = float(value)
                        if score >= 80:
                            cell.fill = SCORE_HIGH
                        elif score >= 60:
                            cell.fill = SCORE_MED
                        else:
                            cell.fill = SCORE_LOW
                    except (ValueError, TypeError):
                        pass

                # compare_columns observations: higher is worse, lower is
                # better. 'same' and empty cells stay unfilled.
                if compare_cols and col_name in compare_cols and value is not None:
                    text = str(value).lower()
                    if "higher" in text:
                        cell.fill = CMP_HIGHER
                    elif "lower" in text:
                        cell.fill = CMP_LOWER


def _auto_width(ws, columns: list, min_width: int = 10, max_width: int = 40):
    """Auto-adjust column widths based on header length."""
    for col_idx, (_, header) in enumerate(columns, 1):
        width = min(max(len(header) + 4, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

def _coalesce_variant_columns(df: pl.DataFrame, column_defs: list) -> pl.DataFrame:
    """Coalesce variant columns created by diagonal concat.

    Handles both recipe-driven column defs (with 'fields' key for variants)
    and legacy DEST_COLUMNS format (multiple entries sharing a header).

    For recipe-driven: entries with ``fields: [col1, col2]`` are coalesced
    into the first present column.

    For legacy: groups DEST_COLUMNS by header and coalesces variants.
    """
    for entry in column_defs:
        if isinstance(entry, dict) and "fields" in entry:
            # Recipe-driven: explicit variant list
            variants = entry["fields"]
            present = [v for v in variants if v in df.columns]
            if len(present) > 1:
                df = df.with_columns(
                    pl.coalesce(present).alias(present[0])
                )
        # Legacy tuple format handled below

    # Legacy: group tuple-format DEST_COLUMNS by header
    header_variants: dict[str, list[str]] = {}
    for entry in column_defs:
        if isinstance(entry, tuple):
            col_name, header = entry
            header_variants.setdefault(header, []).append(col_name)

    for _header, variants in header_variants.items():
        present = [v for v in variants if v in df.columns]
        if len(present) > 1:
            df = df.with_columns(
                pl.coalesce(present).alias(present[0])
            )
    return df


# Backward-compatible alias (used by legacy code paths without recipe-driven columns)
def _coalesce_dest_columns(df):
    return _coalesce_variant_columns(df, DEST_COLUMNS)


def _ensure_reason_columns(df: pl.DataFrame) -> pl.DataFrame:
    """Populate reason_code and rejection fields when absent (Analysis defaults)."""
    if "reason_code" in df.columns:
        return df
    return df.with_columns(
        pl.lit("no_name_match").alias("reason_code"),
        pl.lit(None).cast(pl.String).alias("rejection_step"),
        pl.lit(None).cast(pl.Float64).alias("best_rejected_score"),
    )


def _resolve_columns(df: pl.DataFrame, column_defs: list) -> list:
    """Resolve which columns exist in the DataFrame."""
    available = set(df.columns)
    resolved = []
    seen_headers = set()

    for col_name, header in column_defs:
        if col_name in available and header not in seen_headers:
            resolved.append((col_name, header))
            seen_headers.add(header)

    return resolved


def _build_columns_from_recipe(recipe_columns: list, df: pl.DataFrame) -> list:
    """Build (col_name, header) list from recipe output.columns config.

    Handles both single-field and multi-field (variant) entries.
    For variant entries, returns the first present column after coalescing.
    """
    available = set(df.columns)
    resolved = []

    for entry in recipe_columns:
        header = entry["header"]
        if "fields" in entry:
            # Variant column: use first present (coalesce already ran)
            for f in entry["fields"]:
                if f in available:
                    resolved.append((f, header))
                    break
        elif "field" in entry:
            if entry["field"] in available:
                resolved.append((entry["field"], header))
            else:
                import sys
                print(
                    f'[WARN] Report column "{entry["field"]}" (header: "{header}") '
                    f"not found in data -- skipped",
                    file=sys.stderr,
                )

    return resolved


def _compare_format_columns(recipe: Optional[dict]) -> set:
    """Identifiers of compare_columns outputs, as field names and as headers.

    The merged and non-merged Matched-tab paths label columns differently
    (header vs field), so both spellings are collected and the banding works
    either way.
    """
    if not recipe:
        return set()
    from recipe import compare_output_names

    output_cfg = recipe.get("output", {}) or {}
    fields = set(compare_output_names(output_cfg))
    if not fields:
        return set()

    idents = set(fields)
    for entry in (output_cfg.get("columns") or {}).get("matched") or []:
        if not isinstance(entry, dict):
            continue
        if entry.get("field") in fields and entry.get("header"):
            idents.add(entry["header"])
    return idents


def _known_derived(recipe: Optional[dict]) -> set:
    """Derived/metadata column names the recipe may produce (empty if none)."""
    if not recipe:
        return set()
    from recipe import known_derived_columns
    return known_derived_columns(recipe)


def generate_report(matched_df: pl.DataFrame, unmatched_df: pl.DataFrame | None = None,
                    output_path: str = "", stats: Optional[dict] = None,
                    recipe: Optional[dict] = None,
                    recipe_file: str | None = None,
                    echo_recipe: Optional[dict] = None,
                    merged: bool = False) -> str:
    """Generate the Excel report with Summary, Matched, and Analysis tabs.

    Args:
        matched_df: Matched records from pipeline
        unmatched_df: Unmatched records from pipeline (None for phase output)
        output_path: Path to write the Excel file
        stats: Optional pipeline stats dict
        recipe: Optional recipe dict. When present and output.columns
                is defined, uses recipe-driven column mapping instead
                of hardcoded defaults. Also enables the Summary tab.
        recipe_file: Optional recipe filename for summary tab header
        echo_recipe: Optional recipe dict for the Recipe tab. Defaults to
                recipe. Pass the full multi-phase recipe here when recipe
                is a phase-scoped mini_recipe, so the tab round-trips.
        merged: When True, unmatched source rows are appended into the
                Matched tab (match_step sentinel + is_unmatched flag),
                per the merged output view. Analysis tab is unchanged.

    Returns:
        Path to the generated report
    """
    # openpyxl cannot hold a list cell; the xlsx report renders the same
    # ";"-joined text as the csv/xlsx raw writers.
    matched_df = flatten_list_columns(matched_df)
    unmatched_df = flatten_list_columns(unmatched_df)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    # Determine column source
    recipe_columns = None
    if recipe:
        recipe_columns = recipe.get("output", {}).get("columns", None)

    wb = Workbook()

    # --- Main Report Tab ---
    ws_main = wb.active
    ws_main.title = "Matched"

    merged_active = merged and unmatched_df is not None
    if matched_df.height > 0 or merged_active:
        if recipe_columns and "matched" in recipe_columns:
            # Recipe-driven columns
            matched_df = _coalesce_variant_columns(matched_df, recipe_columns["matched"])
            main_cols = _build_columns_from_recipe(recipe_columns["matched"], matched_df)
        else:
            # Legacy hardcoded columns
            matched_df = _coalesce_dest_columns(matched_df)
            main_cols = _resolve_columns(matched_df, MAIN_REPORT_COLUMNS + DEST_COLUMNS)

        if merged_active:
            # Resolve columns against both frames so an empty matched frame
            # (all-unmatched run) still yields the full merged layout. Falls
            # back to the hardcoded columns when the recipe defines none.
            output_cfg = (recipe or {}).get("output", {})
            derived = _known_derived(recipe)
            merged_cols = None if (recipe_columns and "matched" in recipe_columns) else main_cols
            cols = merged_cols or _resolve_merged_columns(
                matched_df, unmatched_df, output_cfg, derived
            )
            # Keep field names on the frame -- _write_data keys score rounding
            # and the conditional banding off the field, not the header, so
            # renaming here would silently drop both (Issue #95). Headers are
            # applied by _write_headers from main_cols.
            seen: set = set()
            frame_cols = []
            for field, _header in cols:
                if field not in seen:
                    seen.add(field)
                    frame_cols.append((field, field))
            write_df = build_merged_frame(
                matched_df, unmatched_df, output_cfg, derived, columns=frame_cols,
            )
            main_cols = [*cols, ("is_unmatched", "is_unmatched")]
        else:
            write_df = matched_df

        _write_headers(ws_main, main_cols)
        _write_data(ws_main, write_df, main_cols,
                    compare_cols=_compare_format_columns(recipe))
        _auto_width(ws_main, main_cols)

        # Freeze top row
        ws_main.freeze_panes = "A2"
    else:
        ws_main.cell(row=1, column=1, value="No matches found")

    # --- Analysis Tab ---
    ws_analysis = wb.create_sheet("Analysis")

    if unmatched_df is not None and unmatched_df.height > 0:
        unmatched_df = _ensure_reason_columns(unmatched_df)

        if recipe_columns and "analysis" in recipe_columns:
            analysis_cols = _build_columns_from_recipe(recipe_columns["analysis"], unmatched_df)
        else:
            analysis_cols = _resolve_columns(unmatched_df, ANALYSIS_COLUMNS)

        _write_headers(ws_analysis, analysis_cols, header_fill=ANALYSIS_HEADER_FILL)
        _write_data(ws_analysis, unmatched_df, analysis_cols)
        _auto_width(ws_analysis, analysis_cols)

        ws_analysis.freeze_panes = "A2"
    else:
        ws_analysis.cell(row=1, column=1, value="All records matched")

    # --- Summary Tab (first sheet) ---
    if recipe and stats:
        try:
            from summary import write_summary_tab
            ws_summary = wb.create_sheet("Summary", 0)  # Insert at position 0
            write_summary_tab(ws_summary, recipe, stats, matched_df,
                              recipe_file=recipe_file)
        except Exception as exc:
            import sys
            print(f"[WARN] Summary tab generation failed: {exc}", file=sys.stderr)

    # --- Recipe Tab (echo of the resolved recipe, last sheet) ---
    tab_recipe = echo_recipe if echo_recipe is not None else recipe
    if tab_recipe:
        try:
            from recipe_echo import write_recipe_tab
            ws_recipe = wb.create_sheet("Recipe")
            write_recipe_tab(ws_recipe, tab_recipe)
        except Exception as exc:
            import sys
            print(f"[WARN] Recipe tab generation failed: {exc}", file=sys.stderr)

    wb.save(str(out))
    return str(out)


# ---------------------------------------------------------------------------
# Convenience: run full pipeline + generate report
# ---------------------------------------------------------------------------

def run_and_report(recipe_path: str, base_dir: str = ".",
                   output_path: Optional[str] = None) -> str:
    """Run the matching pipeline and generate the Excel report.

    Args:
        recipe_path: Path to recipe YAML/JSON
        base_dir: Base directory for data files
        output_path: Override output path (default from recipe)

    Returns:
        Path to generated report
    """
    from matching import run_pipeline
    from recipe import load_recipe

    recipe = load_recipe(recipe_path)
    result = run_pipeline(recipe, base_dir=base_dir)

    if output_path is None:
        output_path = recipe.get("output", {}).get("path", "output/report.xlsx")

    path = generate_report(
        result["matched"],
        result["unmatched"],
        output_path,
        stats=result["stats"],
        recipe=recipe,
    )

    return path


def sort_by_source_order(df, source_df, key_field: str):
    """Order rows by each key's first appearance in the source input.

    Write-layer only: the matching engine may emit rows in completion order,
    which is not stable run-to-run. Keys absent from the source keep their
    incoming order at the end. Helper columns never reach an artifact.
    """
    if df is None or source_df is None or df.height == 0 or source_df.height == 0:
        return df
    if key_field not in df.columns or key_field not in source_df.columns:
        return df

    order_col, key_col = "_source_order", "_source_order_key"
    while order_col in df.columns or key_col in df.columns:
        order_col, key_col = order_col + "_", key_col + "_"

    order = (
        source_df.select(pl.col(key_field).cast(pl.String).alias(key_col))
        .with_row_index(order_col)
        .group_by(key_col)
        .agg(pl.col(order_col).min())
    )
    return (
        df.with_columns(pl.col(key_field).cast(pl.String).alias(key_col))
        .join(order, on=key_col, how="left")
        .sort(order_col, nulls_last=True, maintain_order=True)
        .drop(order_col, key_col)
    )


def write_raw_data(df, path: str, fmt: str):
    """Write a DataFrame as raw data (csv, xlsx, or parquet)."""
    if fmt == "csv":
        flatten_list_columns(df).write_csv(path)
    elif fmt == "xlsx":
        df = flatten_list_columns(df)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(list(df.columns))
        for row in df.iter_rows():
            ws.append([v for v in row])
        wb.save(path)
    elif fmt == "parquet":
        df.write_parquet(path)
    else:
        raise ValueError(f"Unsupported output format: {fmt}")


def apply_column_mapping(df, output_cfg: dict, key: str = "matched"):
    """Apply an output.columns mapping to a DataFrame.

    Selects and renames columns per the recipe's output.columns[key] config
    ("matched" or "analysis"). Returns the original DataFrame if no mapping
    is configured for that key.
    """
    cols_cfg = output_cfg.get("columns", {})
    col_defs = cols_cfg.get(key)
    if not col_defs:
        return df

    select_cols = []
    rename_map = {}
    for entry in col_defs:
        # Support both 'field' and 'fields' (coalesce)
        field = entry.get("field")
        fields = entry.get("fields", [])
        header = entry.get("header", field)

        if field and field in df.columns:
            select_cols.append(field)
            if header and header != field:
                rename_map[field] = header
        elif field and field not in df.columns and not fields:
            import sys
            print(
                f'[WARN] Output column "{field}" (header: "{header}") '
                f"not found in {key} data -- skipped. "
                f"Available: {', '.join(sorted(df.columns)[:15])}",
                file=sys.stderr,
            )
        elif fields:
            # Coalesce: first non-null field wins
            for f in fields:
                if f in df.columns:
                    select_cols.append(f)
                    if header and header != f:
                        rename_map[f] = header
                    break

    if select_cols:
        df = df.select(select_cols)
    if rename_map:
        df = df.rename(rename_map)
    return df


def write_unmatched_export(unmatched_df, output_cfg: dict, path: str, fmt: str):
    """Write unmatched records as a raw data file (companion to the matched export).

    Columns resolved via output.columns.analysis (recipe-driven, same as the
    Analysis tab). reason_code and rejection fields are populated when absent.
    Zero unmatched rows still writes a header-only file (deterministic for DW
    imports). No-ops (returns None) only when there is no unmatched frame.
    """
    if unmatched_df is None:
        return None
    from recipe import output_computed_columns
    # decision_record/compare columns are matched-view only (they may ride on
    # the unmatched frame purely to reach the merged view).
    stowaways = [c for c in output_computed_columns(output_cfg) if c in unmatched_df.columns]
    if stowaways:
        unmatched_df = unmatched_df.drop(stowaways)
    unmatched_df = _ensure_reason_columns(unmatched_df)
    export_df = apply_column_mapping(unmatched_df, output_cfg, key="analysis")
    write_raw_data(export_df, path, fmt)
    return path


def _resolve_merged_columns(matched_df, unmatched_df, output_cfg: dict,
                            derived: set | None = None) -> list:
    """Ordered (field, header) list for the merged raw export."""
    derived = derived or set()
    col_defs = (output_cfg.get("columns") or {}).get("matched")
    if not col_defs:
        base = list(matched_df.columns) or list(unmatched_df.columns)
        return [(c, c) for c in base]

    resolved = []
    for entry in col_defs:
        header = entry.get("header")
        field = None
        if "field" in entry:
            field = entry["field"]
        elif "fields" in entry:
            for f in entry["fields"]:
                if f in matched_df.columns or f in unmatched_df.columns:
                    field = f
                    break
            if field is None and entry["fields"]:
                field = entry["fields"][0]
        if field is None:
            continue
        if (field in matched_df.columns or field in unmatched_df.columns
                or field == "match_step" or field in derived):
            resolved.append((field, header or field))
    return resolved


def build_merged_frame(matched_df, unmatched_df, output_cfg: dict,
                       derived: set | None = None, columns: list | None = None):
    """Matched rows + unmatched source rows as one presentation frame."""
    cols = columns or _resolve_merged_columns(
        matched_df, unmatched_df, output_cfg, derived
    )

    def _side(df, is_unmatched: bool):
        exprs = []
        for field, header in cols:
            if is_unmatched and field == "match_step":
                expr = pl.lit("unmatched")
            elif field in df.columns:
                expr = pl.col(field)
            else:
                expr = pl.lit(None)
            exprs.append(expr.alias(header))
        side = df.select(exprs).with_columns(pl.lit(is_unmatched).alias("is_unmatched"))
        # A 0-column, 0-row frame broadcasts literals to a phantom row; drop it.
        return side.clear() if df.height == 0 else side

    matched_side = _side(matched_df, False)
    if unmatched_df is None:
        return matched_side
    unmatched_side = _side(unmatched_df, True)
    return pl.concat([matched_side, unmatched_side], how="vertical_relaxed")


# ---------------------------------------------------------------------------
# Presentation-layer computed columns (Issue #93)
# ---------------------------------------------------------------------------

def _blank_as_null(df, col: str) -> pl.Expr:
    """Column as trimmed String, with absent/empty/whitespace-only as null."""
    if col not in df.columns:
        return pl.lit(None, pl.String)
    trimmed = pl.col(col).cast(pl.String).str.strip_chars()
    return pl.when(trimmed == "").then(None).otherwise(trimmed)


def _parsed_pair(df, col: str, strip_prefix: str):
    """(stripped string, Float64) for a column, sharing the tie-breaker helper.

    Whitespace-strip, apply strip_prefix, then parse. Float64 (not the
    engine's Int64) so decimal identifiers and amounts compare numerically;
    only values that fail the float parse fall back to text. Nullity is
    decided on the value *before* stripping, so a prefix-only value like
    "AB" stays a present (empty-after-strip) value rather than becoming null.
    """
    from normalize import strip_prefix_expr

    present = _blank_as_null(df, col)
    # "none" disables stripping for these features; the engine has no such
    # spelling, so the mapping happens here rather than in the shared helper.
    resolved = "" if strip_prefix == "none" else strip_prefix
    stripped = strip_prefix_expr(present, resolved).str.strip_chars()
    return stripped, stripped.cast(pl.Float64, strict=False)


def _decision_record_exprs(df, cfg: dict) -> list:
    """<write_to> / <write_to>_src expressions for one decision_record."""
    candidates = list(cfg.get("candidates") or [])
    write_to = cfg["write_to"]
    select = cfg.get("select", "first")
    strip_prefix = cfg.get("strip_prefix", "alpha")

    values = [_blank_as_null(df, c) for c in candidates]
    keys = [_parsed_pair(df, c, strip_prefix)[1] for c in candidates]

    def _first_index(conds):
        """Index of the earliest candidate satisfying its condition."""
        expr = pl.lit(None, pl.Int32)
        for i, cond in reversed(list(enumerate(conds))):
            expr = pl.when(cond).then(pl.lit(i, pl.Int32)).otherwise(expr)
        return expr

    populated = [v.is_not_null() for v in values]
    list_order_idx = _first_index(populated)

    if select in ("min", "max"):
        # Only populated candidates compete; a candidate that will not parse
        # cannot win on value. Ties and an all-unparseable row fall back to
        # list order, matching the tie-breaker's own preference rule.
        eligible = [
            pl.when(p).then(k).otherwise(None)
            for p, k in zip(populated, keys, strict=True)
        ]
        target = (
            pl.min_horizontal(eligible) if select == "min"
            else pl.max_horizontal(eligible)
        )
        best = _first_index([e.is_not_null() & (e == target) for e in eligible])
        idx = pl.coalesce([best, list_order_idx])
    else:
        idx = list_order_idx

    value_expr = pl.lit(None, pl.String)
    src_expr = pl.lit(None, pl.String)
    for i, (name, val) in reversed(list(enumerate(zip(candidates, values, strict=True)))):
        hit = idx == i
        value_expr = pl.when(hit).then(val).otherwise(value_expr)
        src_expr = pl.when(hit).then(pl.lit(name)).otherwise(src_expr)

    return [
        value_expr.alias(write_to),
        src_expr.alias(f"{write_to}_src"),
    ]


def _compare_expr(df, entry: dict) -> pl.Expr:
    """higher/lower/same observation for one compare_columns entry.

    Never fails on data. Both sides parsing as integers gives a numeric
    comparison; anything else falls back to lexicographic on the stripped
    strings. A null on either side yields an empty cell.
    """
    strip_prefix = entry.get("strip_prefix", "alpha")
    left_s, left_n = _parsed_pair(df, entry["left"], strip_prefix)
    right_s, right_n = _parsed_pair(df, entry["right"], strip_prefix)

    def _verdict(left, right):
        return (
            pl.when(left > right).then(pl.lit("higher"))
            .when(left < right).then(pl.lit("lower"))
            .otherwise(pl.lit("same"))
        )

    return (
        pl.when(left_s.is_null() | right_s.is_null())
        .then(pl.lit(None, pl.String))
        .when(left_n.is_not_null() & right_n.is_not_null())
        .then(_verdict(left_n, right_n))
        .otherwise(_verdict(left_s, right_s))
        .alias(entry["output"])
    )


def _group_hit(df, group: dict) -> pl.Expr:
    """Boolean: does this group apply to each row?

    Matching runs on raw source values only -- no tier/normalization. A group
    applies iff any match_column matches ``regex`` or equals (trimmed,
    case-insensitive) an entry in ``values``, and no match_column matches
    ``exclude_regex``. An absent match_column contributes nothing rather than
    failing, so the same expression works on the unmatched frame.
    """
    present = [c for c in (group.get("match_columns") or []) if c in df.columns]
    if not present:
        return pl.lit(False)

    raw = [pl.col(c).cast(pl.String) for c in present]
    hits = []
    if group.get("regex"):
        hits.extend(r.str.contains(group["regex"]) for r in raw)
    values = group.get("values") or []
    if values:
        wanted = [str(v).strip().casefold() for v in values]
        hits.extend(r.str.strip_chars().str.to_lowercase().is_in(wanted) for r in raw)
    included = pl.any_horizontal(hits).fill_null(False)

    if group.get("exclude_regex"):
        excluded = pl.any_horizontal(
            [r.str.contains(group["exclude_regex"]) for r in raw]
        ).fill_null(False)
        return included & ~excluded
    return included


def _groups_expr(df, cfg: dict, group_defs: list) -> pl.Expr:
    """The ``group`` tag column for one output.groups config.

    all_match  -> List(String), every matching name sorted, [] for no match.
    first_match -> String, the first matching group in file order, null for
    no match. File order is priority; there is no scoring rule.
    """
    from recipe import GROUP_COLUMN

    hits = [(g.get("group_name"), _group_hit(df, g)) for g in group_defs]

    if cfg.get("mode", "all_match") == "first_match":
        expr = pl.lit(None, pl.String)
        for name, hit in reversed(hits):
            expr = pl.when(hit).then(pl.lit(name)).otherwise(expr)
        return expr.alias(GROUP_COLUMN)

    tagged = [
        pl.when(hit).then(pl.lit(name)).otherwise(pl.lit(None, pl.String))
        for name, hit in hits
    ]
    return (
        pl.concat_list(tagged).list.drop_nulls().list.sort().alias(GROUP_COLUMN)
    )


def flatten_list_columns(df, sep: str = ";"):
    """Join List(String) columns into scalars for csv/xlsx serialization.

    parquet keeps the native list; the text formats cannot hold one and
    openpyxl rejects a list cell outright.
    """
    if df is None:
        return df
    list_cols = [c for c, dt in zip(df.columns, df.dtypes, strict=True)
                 if isinstance(dt, pl.List)]
    if not list_cols:
        return df
    return df.with_columns(
        pl.col(c).cast(pl.List(pl.String)).list.join(sep) for c in list_cols
    )


def _matched_column_fields(output_cfg: dict) -> set:
    """Field names referenced by output.columns.matched."""
    referenced = set()
    for entry in (output_cfg.get("columns") or {}).get("matched") or []:
        if not isinstance(entry, dict):
            continue
        if "field" in entry:
            referenced.add(entry["field"])
        referenced.update(entry.get("fields") or [])
    return referenced


def apply_output_computations(df, output_cfg: dict, group_defs: list | None = None):
    """Attach decision_record / compare_columns / groups columns to a frame.

    Runs on the matched and unmatched frames alike, so the merged view picks
    both up with no merged-view code: a candidate absent from the unmatched
    frame is null there and the decision falls through to a source-side
    candidate.

    Computed columns are pruned unless ``output.columns.matched`` names them --
    these features never auto-append a column to an artifact.
    """
    from recipe import output_computed_columns

    computed = output_computed_columns(output_cfg)
    if df is None or not computed or df.height == 0:
        return df

    # groups lands in its own earlier pass: it reads only raw source values,
    # and a follow-up's scalar emit columns must be materialized before the
    # decision_record exprs can name them as candidates.
    if group_defs:
        df = df.with_columns(_groups_expr(df, output_cfg["groups"], group_defs))

    exprs = []
    if (output_cfg.get("decision_record") or {}).get("write_to"):
        exprs.extend(_decision_record_exprs(df, output_cfg["decision_record"]))
    for entry in output_cfg.get("compare_columns") or []:
        exprs.append(_compare_expr(df, entry))

    df = df.with_columns(exprs)

    referenced = _matched_column_fields(output_cfg)
    drop = [c for c in computed if c not in referenced and c in df.columns]
    return df.drop(drop) if drop else df


def enrich_join(source_df, matched_df, enrich_key: str):
    """Left-join matched results onto a source dataset for enriched output.

    Every source row appears in output. Matched rows get enrichment
    columns (columns in matched_df not in source_df) populated;
    unmatched rows get null.

    Returns (enriched_df, matched_count).
    Raises ValueError if enrich_key is missing from either DataFrame.
    """
    if enrich_key not in source_df.columns:
        raise ValueError(
            f"enrich_key '{enrich_key}' not found in enrichment source. "
            f"Available: {source_df.columns[:10]}"
        )

    matched_count = matched_df.height

    if matched_df.height == 0:
        return source_df, 0

    if enrich_key not in matched_df.columns:
        raise ValueError(
            f"enrich_key '{enrich_key}' not found in matched results. "
            f"Available: {matched_df.columns[:10]}"
        )

    # Only bring in columns that don't already exist in the source
    enrichment_cols = [
        c for c in matched_df.columns if c not in source_df.columns
    ]
    if not enrichment_cols:
        return source_df, matched_count

    enrich_df = matched_df.select([enrich_key] + enrichment_cols)
    enrich_df = enrich_df.unique(subset=[enrich_key], keep="first")

    # maintain_order: enriched output is the source in its own input order.
    enriched = source_df.join(
        enrich_df, on=enrich_key, how="left", maintain_order="left"
    )
    return enriched, matched_count
