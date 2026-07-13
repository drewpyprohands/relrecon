"""Echo the resolved recipe into report outputs (xlsx Recipe tab, md section)."""

from __future__ import annotations

import json

try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

# "#" lines are YAML comments, so they survive a round-trip parse.
_HEADER_LINES = [
    "# Resolved recipe -- full config incl. exclusions, derived columns, tie-breakers.",
    "# Round-trips: parse this text as YAML to reconstruct an equivalent recipe.",
]


def serialize_recipe(recipe: dict) -> str:
    """Serialize a resolved recipe to round-trippable text (YAML, else JSON)."""
    if YAML_AVAILABLE:
        body = yaml.safe_dump(
            recipe, sort_keys=False, default_flow_style=False, allow_unicode=True
        )
        return "\n".join(_HEADER_LINES) + "\n" + body
    # JSON fallback omits the "#" header -- it would break json.loads.
    return json.dumps(recipe, indent=2, ensure_ascii=False)


def parse_serialized_recipe(text: str) -> dict:
    """Parse serialized recipe text back into a dict (YAML superset handles JSON)."""
    if YAML_AVAILABLE:
        return yaml.safe_load(text)
    return json.loads(text)


def write_recipe_tab(ws, recipe: dict) -> None:
    """Write the resolved recipe into a worksheet, one line per row in column A."""
    from openpyxl.styles import Font

    mono = Font(name="Consolas", size=10)
    text = serialize_recipe(recipe)
    for i, line in enumerate(text.split("\n"), start=1):
        cell = ws.cell(row=i, column=1, value=line if line else None)
        cell.font = mono
    ws.column_dimensions["A"].width = 100


def read_recipe_tab(ws) -> dict:
    """Reconstruct the recipe from a worksheet written by write_recipe_tab."""
    lines = []
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        lines.append("" if val is None else str(val))
    return parse_serialized_recipe("\n".join(lines))


def recipe_md_section(recipe: dict) -> str:
    """Return a markdown section echoing the resolved recipe in a fenced block."""
    fence = "yaml" if YAML_AVAILABLE else "json"
    body = serialize_recipe(recipe)
    return (
        "## Recipe (resolved)\n\n"
        "Full config incl. exclusions, derived columns, and tie-breakers. "
        "This block round-trips to an equivalent recipe.\n\n"
        f"```{fence}\n{body.rstrip()}\n```"
    )
