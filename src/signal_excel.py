"""
Signal analysis Excel report generator.

Produces a multi-sheet Excel workbook from analyze_dataset() results,
matching the formatting conventions of report.py (headers, borders, widths).

Sheets:
- Summary: top tokens preview + data quality side-by-side
- TopTokens: full token detail (topk, bigram, trigram, singleton) per column/tier
- Alias: detected punctuation/case variant groups
- NearDuplicates: token pairs with high edit similarity (probable typos)
- TokenProfile: position frequency, length distribution, numeric ratio
- Unicode: character range profiles per column
"""

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Styling (matches report.py conventions)
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
SECTION_FONT = Font(bold=True, size=12)
SUBHEADER_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Frequency coloring thresholds
FREQ_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red -- very frequent
FREQ_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Yellow -- moderate
FREQ_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Green -- low freq

# Similarity coloring
SIM_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Red -- very similar (likely typo)
SIM_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")    # Yellow -- moderate


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _style_header(ws, row: int, col: int, value: str,
                  font=HEADER_FONT, fill=HEADER_FILL):
    """Write a styled header cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.fill = fill
    cell.alignment = HEADER_ALIGNMENT
    cell.border = THIN_BORDER
    return cell


def _write_cell(ws, row: int, col: int, value, border=True):
    """Write a data cell with optional border."""
    cell = ws.cell(row=row, column=col, value=value)
    if border:
        cell.border = THIN_BORDER
    return cell


def _auto_width(ws, col_widths: dict):
    """Set column widths from a {col_letter: width} dict."""
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width


def _freq_fill(freq: float) -> Optional[PatternFill]:
    """Return a fill color based on frequency value."""
    if freq >= 0.20:
        return FREQ_HIGH
    if freq >= 0.10:
        return FREQ_MED
    return None


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary_sheet(ws, results: dict, top_n: int = 25):
    """Build the Summary sheet with top tokens preview + data quality."""

    # -- Left side: Top Tokens preview --
    _write_cell(ws, 1, 1, "Top Tokens", border=False).font = SECTION_FONT
    _write_cell(ws, 2, 1,
                f"Top {top_n} tokens by count across all columns/tiers "
                "(see TopTokens sheet for full detail)", border=False)

    # Headers at row 4
    token_headers = ["colName", "token", "dataTier", "uniqueColCount"]
    for i, h in enumerate(token_headers, 1):
        _style_header(ws, 4, i, h)

    # Collect ALL tokens across columns, sort by count, take top N
    all_tokens = []
    for col_name, col_data in results.get("columns", {}).items():
        for tier_key, tier_label in [("top_tokens_raw", "raw"),
                                     ("top_tokens_clean", "clean")]:
            for token, count in col_data.get(tier_key, []):
                all_tokens.append((col_name, token, tier_label, count))

    # Sort by count descending, take top N
    all_tokens.sort(key=lambda x: -x[3])
    preview = all_tokens[:top_n]

    row = 5
    for col_name, token, tier_label, count in preview:
        _write_cell(ws, row, 1, col_name)
        _write_cell(ws, row, 2, token)
        _write_cell(ws, row, 3, tier_label)
        _write_cell(ws, row, 4, count)
        row += 1

    # -- Right side: Data Quality --
    dq_start_col = 7
    _write_cell(ws, 1, dq_start_col, "Data Quality", border=False).font = SECTION_FONT

    dq_headers = ["colName", "detectedType", "rows", "null%", "unique%",
                  "duplicates", "numeric%"]
    for i, h in enumerate(dq_headers):
        _style_header(ws, 4, dq_start_col + i, h)

    quality = results.get("data_quality", {})
    dq_row = 5
    for col_name, q in quality.items():
        col_type = "unknown"
        if col_name in results.get("columns", {}):
            col_type = results["columns"][col_name].get("detected_type", "unknown")
        _write_cell(ws, dq_row, dq_start_col, col_name)
        _write_cell(ws, dq_row, dq_start_col + 1, col_type)
        _write_cell(ws, dq_row, dq_start_col + 2, q.get("total_rows", 0))
        _write_cell(ws, dq_row, dq_start_col + 3, q.get("null_pct", 0) / 100)
        _write_cell(ws, dq_row, dq_start_col + 4, q.get("unique_pct", 0) / 100)
        _write_cell(ws, dq_row, dq_start_col + 5, q.get("duplicate_count", 0))
        _write_cell(ws, dq_row, dq_start_col + 6, q.get("numeric_token_pct", 0) / 100)
        # Format percentages
        for offset in (3, 4, 6):
            ws.cell(row=dq_row, column=dq_start_col + offset).number_format = '0.0%'
        dq_row += 1

    # Column widths
    _auto_width(ws, {
        "A": 18, "B": 14, "C": 10, "D": 17,
        "F": 3,  # spacer
        "G": 18, "H": 15, "I": 8, "J": 8, "K": 10, "L": 12, "M": 10,
    })


def _build_top_tokens_sheet(ws, results: dict, top_n: Optional[int] = None):
    """Build the TopTokens sheet with full detail including singletons.

    Each row: columnName, signalType (topk/bigram/trigram/singleton),
    dataTier, token, rows, freq%
    """
    headers = ["columnName", "signalType", "dataTier", "token", "rows", "freq % per col"]
    for i, h in enumerate(headers, 1):
        _style_header(ws, 1, i, h)

    row = 2
    for col_name, col_data in results.get("columns", {}).items():
        # Get total rows for frequency calculation
        quality = results.get("data_quality", {}).get(col_name, {})
        total_rows = quality.get("non_null", quality.get("total_rows", 1)) or 1

        # Top-K tokens
        for tier_key, tier_label in [("top_tokens_raw", "raw"),
                                     ("top_tokens_clean", "clean")]:
            tokens = col_data.get(tier_key, [])
            if top_n:
                tokens = tokens[:top_n]
            for token, count in tokens:
                freq = count / total_rows if total_rows > 0 else 0
                _write_cell(ws, row, 1, col_name)
                _write_cell(ws, row, 2, "topk")
                _write_cell(ws, row, 3, tier_label)
                _write_cell(ws, row, 4, token)
                _write_cell(ws, row, 5, count)
                cell = _write_cell(ws, row, 6, freq)
                cell.number_format = '0.00'
                fill = _freq_fill(freq)
                if fill:
                    cell.fill = fill
                row += 1

        # Bigrams
        for tier_key, tier_label in [("bigrams_raw", "raw"),
                                     ("bigrams_clean", "clean")]:
            bigrams = col_data.get(tier_key, [])
            if top_n:
                bigrams = bigrams[:top_n]
            for token, count in bigrams:
                freq = count / total_rows if total_rows > 0 else 0
                _write_cell(ws, row, 1, col_name)
                _write_cell(ws, row, 2, "bigram")
                _write_cell(ws, row, 3, tier_label)
                _write_cell(ws, row, 4, token)
                _write_cell(ws, row, 5, count)
                cell = _write_cell(ws, row, 6, freq)
                cell.number_format = '0.00'
                fill = _freq_fill(freq)
                if fill:
                    cell.fill = fill
                row += 1

        # Trigrams
        for tier_key, tier_label in [("trigrams_raw", "raw"),
                                     ("trigrams_clean", "clean")]:
            trigrams = col_data.get(tier_key, [])
            if top_n:
                trigrams = trigrams[:top_n]
            for token, count in trigrams:
                freq = count / total_rows if total_rows > 0 else 0
                _write_cell(ws, row, 1, col_name)
                _write_cell(ws, row, 2, "trigram")
                _write_cell(ws, row, 3, tier_label)
                _write_cell(ws, row, 4, token)
                _write_cell(ws, row, 5, count)
                cell = _write_cell(ws, row, 6, freq)
                cell.number_format = '0.00'
                fill = _freq_fill(freq)
                if fill:
                    cell.fill = fill
                row += 1

        # Singletons
        singletons = col_data.get("singletons", [])
        if top_n:
            singletons = singletons[:top_n]
        for token, count in singletons:
            freq = count / total_rows if total_rows > 0 else 0
            _write_cell(ws, row, 1, col_name)
            _write_cell(ws, row, 2, "singleton")
            _write_cell(ws, row, 3, "clean")
            _write_cell(ws, row, 4, token)
            _write_cell(ws, row, 5, count)
            cell = _write_cell(ws, row, 6, freq)
            cell.number_format = '0.00'
            row += 1

    _auto_width(ws, {
        "A": 18, "B": 10, "C": 10, "D": 20, "E": 8, "F": 16,
    })


def _build_alias_sheet(ws, results: dict):
    """Build the Alias sheet with variant groups."""
    headers = ["canonical", "variant", "count", "totalGroupCount"]
    for i, h in enumerate(headers, 1):
        _style_header(ws, 1, i, h)

    row = 2
    for _col_name, col_data in results.get("columns", {}).items():
        aliases = col_data.get("suggested_aliases", [])
        for ag in aliases:
            canonical = ag.get("canonical", "")
            total = ag.get("total_count", 0)
            for variant in ag.get("variants", []):
                _write_cell(ws, row, 1, canonical)
                _write_cell(ws, row, 2, variant.get("raw", ""))
                _write_cell(ws, row, 3, variant.get("count", 0))
                _write_cell(ws, row, 4, total)
                row += 1

    # Also include aggregated aliases
    agg = results.get("aggregated_aliases", {})
    if agg and row == 2:
        for variant_clean, canonical in agg.items():
            _write_cell(ws, row, 1, canonical)
            _write_cell(ws, row, 2, variant_clean)
            row += 1

    if row == 2:
        ws.cell(row=2, column=1,
                value="No punctuation/case variants detected. "
                      "Semantic aliases (blvd/boulevard, st/street) must be added manually "
                      "to aliases.json.")

    _auto_width(ws, {"A": 16, "B": 20, "C": 10, "D": 18})


def _build_near_duplicates_sheet(ws, results: dict):
    """Build the NearDuplicates sheet with edit-distance token pairs."""
    headers = ["columnName", "token1", "token2", "similarity%",
               "count1", "count2"]
    for i, h in enumerate(headers, 1):
        _style_header(ws, 1, i, h)

    row = 2
    for col_name, col_data in results.get("columns", {}).items():
        near_dupes = col_data.get("near_duplicates", [])
        for nd in near_dupes:
            _write_cell(ws, row, 1, col_name)
            _write_cell(ws, row, 2, nd["token1"])
            _write_cell(ws, row, 3, nd["token2"])
            sim_cell = _write_cell(ws, row, 4, nd["similarity"])
            if nd["similarity"] >= 95:
                sim_cell.fill = SIM_HIGH
            elif nd["similarity"] >= 90:
                sim_cell.fill = SIM_MED
            _write_cell(ws, row, 5, nd["count1"])
            _write_cell(ws, row, 6, nd["count2"])
            row += 1

    if row == 2:
        ws.cell(row=2, column=1,
                value="No near-duplicate tokens detected above similarity threshold.")

    _auto_width(ws, {"A": 18, "B": 20, "C": 20, "D": 14, "E": 10, "F": 10})


def _build_token_profile_sheet(ws, results: dict, top_n: Optional[int] = None):
    """Build the TokenProfile sheet with position, length and numeric data."""

    row = 1
    _write_cell(ws, row, 1, "Token Position Frequency", border=False).font = SECTION_FONT
    row += 1
    _write_cell(ws, row, 1,
                "Where tokens appear: first word, last word, or middle",
                border=False)
    row += 1

    # Position frequency headers
    pos_headers = ["columnName", "position", "token", "count"]
    for i, h in enumerate(pos_headers, 1):
        _style_header(ws, row, i, h)
    row += 1

    for col_name, col_data in results.get("columns", {}).items():
        positions = col_data.get("token_positions", {})
        for pos_name in ("first", "last", "middle"):
            pos_data = positions.get(pos_name, [])
            if top_n:
                pos_data = pos_data[:top_n]
            for token, count in pos_data:
                _write_cell(ws, row, 1, col_name)
                _write_cell(ws, row, 2, pos_name)
                _write_cell(ws, row, 3, token)
                _write_cell(ws, row, 4, count)
                row += 1

    # Spacer
    row += 2
    _write_cell(ws, row, 1, "Token Length Distribution", border=False).font = SECTION_FONT
    row += 1
    _write_cell(ws, row, 1,
                "Character count per token (clean tier)", border=False)
    row += 1

    # Length stats headers
    len_headers = ["columnName", "min", "max", "mean", "median"]
    for i, h in enumerate(len_headers, 1):
        _style_header(ws, row, i, h)
    row += 1

    for col_name, col_data in results.get("columns", {}).items():
        tl = col_data.get("token_lengths", {})
        if tl:
            _write_cell(ws, row, 1, col_name)
            _write_cell(ws, row, 2, tl.get("min", 0))
            _write_cell(ws, row, 3, tl.get("max", 0))
            _write_cell(ws, row, 4, tl.get("mean", 0))
            _write_cell(ws, row, 5, tl.get("median", 0))
            row += 1

    # Length histogram
    row += 1
    hist_headers = ["columnName", "tokenLength", "count"]
    for i, h in enumerate(hist_headers, 1):
        _style_header(ws, row, i, h)
    row += 1

    for col_name, col_data in results.get("columns", {}).items():
        tl = col_data.get("token_lengths", {})
        histogram = tl.get("histogram", [])
        if top_n:
            histogram = histogram[:top_n]
        for length, count in histogram:
            _write_cell(ws, row, 1, col_name)
            _write_cell(ws, row, 2, length)
            _write_cell(ws, row, 3, count)
            row += 1

    # Spacer + Numeric ratio
    row += 2
    _write_cell(ws, row, 1, "Numeric Token Ratio", border=False).font = SECTION_FONT
    row += 1
    nr_headers = ["columnName", "totalTokens", "alpha", "numeric",
                  "mixed", "numeric%"]
    for i, h in enumerate(nr_headers, 1):
        _style_header(ws, row, i, h)
    row += 1

    for col_name, col_data in results.get("columns", {}).items():
        nr = col_data.get("numeric_ratio", {})
        if nr:
            _write_cell(ws, row, 1, col_name)
            _write_cell(ws, row, 2, nr.get("total_tokens", 0))
            _write_cell(ws, row, 3, nr.get("alpha", 0))
            _write_cell(ws, row, 4, nr.get("numeric", 0))
            _write_cell(ws, row, 5, nr.get("mixed", 0))
            cell = _write_cell(ws, row, 6, nr.get("numeric_pct", 0) / 100)
            cell.number_format = '0.0%'
            row += 1

    _auto_width(ws, {
        "A": 18, "B": 14, "C": 10, "D": 10, "E": 10, "F": 12,
    })


def _build_unicode_sheet(ws, results: dict, top_n: Optional[int] = None):
    """Build the Unicode sheet with character range profiles."""
    headers = ["columnName", "range", "characters", "pct",
               "cellsWithUnknown", "mixedScriptCells"]
    for i, h in enumerate(headers, 1):
        _style_header(ws, 1, i, h)

    row = 2
    for col_name, col_data in results.get("columns", {}).items():
        up = col_data.get("unicode_profile")
        if not up:
            continue

        bt = up.get("bucket_totals", {})
        total_chars = sum(bt.values()) or 1
        unknown = up.get("cells_with_unknown", 0)
        mixed = up.get("mixed_script_cells", 0)

        sorted_buckets = sorted(bt.items(), key=lambda x: -x[1])
        if top_n:
            sorted_buckets = sorted_buckets[:top_n]

        first_row = True
        for rng, count in sorted_buckets:
            _write_cell(ws, row, 1, col_name)
            _write_cell(ws, row, 2, rng)
            _write_cell(ws, row, 3, count)
            pct = count / total_chars
            cell = _write_cell(ws, row, 4, pct)
            cell.number_format = '0.0%'
            if first_row:
                _write_cell(ws, row, 5, unknown)
                _write_cell(ws, row, 6, mixed)
                first_row = False
            row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="No unicode profiles generated.")

    _auto_width(ws, {"A": 18, "B": 20, "C": 14, "D": 8, "E": 18, "F": 18})


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_signal_excel(results: dict, output_path: str,
                          top_n: Optional[int] = None,
                          summary_top_n: int = 25) -> str:
    """Generate an Excel workbook from signal analysis results.

    Args:
        results: Output from analyze_dataset()
        output_path: Path for the .xlsx file
        top_n: Max items per section in detail sheets (None = all)
        summary_top_n: Max preview rows in Summary sheet (default 25)

    Returns:
        Path to the written file
    """
    wb = Workbook()

    # Summary sheet (default first sheet)
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _build_summary_sheet(ws_summary, results, top_n=summary_top_n)

    # TopTokens sheet
    ws_tokens = wb.create_sheet("TopTokens")
    _build_top_tokens_sheet(ws_tokens, results, top_n=top_n)

    # Alias sheet
    ws_alias = wb.create_sheet("Alias")
    _build_alias_sheet(ws_alias, results)

    # NearDuplicates sheet
    ws_dupes = wb.create_sheet("NearDuplicates")
    _build_near_duplicates_sheet(ws_dupes, results)

    # TokenProfile sheet
    ws_profile = wb.create_sheet("TokenProfile")
    _build_token_profile_sheet(ws_profile, results, top_n=top_n)

    # Unicode sheet
    ws_unicode = wb.create_sheet("Unicode")
    _build_unicode_sheet(ws_unicode, results, top_n=top_n)

    # Write
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)
