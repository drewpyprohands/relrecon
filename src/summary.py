"""
Run summary generator.

Produces a plain-English summary of a pipeline run by combining the
recipe config (what was configured) with the pipeline stats (what
actually happened). Output as markdown string or Excel "Summary" tab.
"""

from __future__ import annotations

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import polars as pl


def _get_all_steps(recipe: dict) -> list:
    """Extract all steps from a recipe, handling both single and multi-phase."""
    if "steps" in recipe and recipe["steps"]:
        return recipe["steps"]
    if "phases" in recipe:
        all_steps = []
        for phase in recipe["phases"]:
            for step in phase.get("steps", []):
                all_steps.append(step)
        return all_steps
    return []


def _get_all_populations(recipe: dict) -> dict:
    """Extract all population configs, handling both single and multi-phase."""
    if "populations" in recipe:
        return recipe["populations"]
    if "phases" in recipe:
        all_pops = {}
        for phase in recipe["phases"]:
            all_pops.update(phase.get("populations", {}))
        return all_pops
    return {}


def _describe_filters(pop_cfg: dict) -> str:
    """Describe population filters in plain English."""
    filters = pop_cfg.get("filter", [])
    if not filters:
        return "everything remaining"

    parts = []
    for f in filters:
        field = f.get("field", "?")
        op = f.get("op", "?")
        val = f.get("value", f.get("values", "?"))

        if op == "starts_with":
            parts.append(f'{field} starts with "{val}"')
        elif op == "not_starts_with":
            parts.append(f'{field} does not start with "{val}"')
        elif op == "eq":
            parts.append(f'{field} is "{val}"')
        elif op == "neq":
            parts.append(f'{field} is not "{val}"')
        elif op == "contains":
            parts.append(f'{field} contains "{val}"')
        elif op == "contains_any":
            quoted = [f'"{v}"' for v in val] if isinstance(val, list) else [f'"{val}"']
            parts.append(f'{field} contains any of {", ".join(quoted)}')
        else:
            parts.append(f"{field} {op} {val}")

    join_mode = "and"
    for f in filters:
        if "join" in f:
            join_mode = f["join"]
            break

    return f" {join_mode} ".join(parts)


def _describe_step_enhanced(step: dict, step_index: int, matched_count: int,
                           total_matched: int, total_source: int,
                           cumulative_matched: int) -> dict:
    """Extract full step info into a flat dict for rendering."""
    mf = step.get("match_fields", [{}])[0]
    method = mf.get("method", "?")
    threshold = mf.get("threshold", 100 if method == "exact" else "?")

    addr = step.get("address_support", {})
    addr_threshold = addr.get("threshold", None)
    addr_tiers = addr.get("tiers", [])

    # Date filter from date_gate or filters
    date_desc = "-"
    dg = step.get("date_gate")
    if dg:
        applies = dg.get("applies_to", "")
        suffix = f" ({applies})" if applies else ""
        date_desc = f'{dg["field"].capitalize()} < {dg["max_age_years"]}yr{suffix}'
    else:
        for f in step.get("filters", []):
            if f.get("op") == "max_age_years":
                applies = f.get("applies_to", "")
                suffix = f" ({applies})" if applies else ""
                date_desc = f'{f["field"].capitalize()} < {f["value"]}yr{suffix}'
                break

    # Other conditions
    conditions = []
    if addr.get("require_street_match"):
        conditions.append("require_street_match")
    weights = addr.get("weights", {})
    sw = weights.get("street_name", weights.get("street_weight", None))
    if sw is not None and sw != 0.6:
        conditions.append(f"street_weight={sw}")
    scorer = mf.get("scorer")
    if scorer and scorer != "token_sort_ratio":
        conditions.append(f"scorer={scorer}")
    other_cond = ", ".join(conditions) if conditions else "-"

    # Percentages
    pct_matched = round(matched_count / total_matched * 100, 1) if total_matched > 0 else 0.0
    # % of leftovers = what fraction of the remaining pool this step matched
    pool_before = total_source - (cumulative_matched - matched_count)
    pct_leftovers = round(matched_count / pool_before * 100, 1) if pool_before > 0 else 0.0

    return {
        "step": step_index,
        "source_pop": step.get("source", "?"),
        "source_col": mf.get("source", "?"),
        "dest_pop": step.get("destination", "?"),
        "dest_col": mf.get("destination", "?"),
        "method": method.capitalize(),
        "data_tier": ", ".join(mf.get("tiers", [])) or "-",
        "name_threshold": str(threshold) if method != "exact" and isinstance(threshold, (int, float)) else "-",
        "addr_threshold": str(addr_threshold) if isinstance(addr_threshold, (int, float)) else "-",
        "addr_tier": ", ".join(addr_tiers) if addr_tiers else "-",
        "date_filter": date_desc,
        "other_conditions": other_cond,
        "matched": matched_count,
        "pct_matched": f"{pct_matched}%",
        "pct_leftovers": f"{pct_leftovers}%",
    }


def _format_timing(timing: dict) -> str:
    """Format timing dict into a readable string."""
    phases = [("load", "Load"), ("setup", "Setup"), ("match", "Match"), ("resolve", "Resolve")]
    parts = [f"{label} {timing[k]:.2f}s" for k, label in phases if k in timing]
    total = sum(timing.get(k, 0) for k, _ in phases)
    parts.append(f"Total {total:.2f}s")
    return " | ".join(parts)


def generate_summary(recipe: dict, stats: dict, matched_df: pl.DataFrame,
                     timing: dict | None = None,
                     mermaid: str = "default",
                     recipe_file: str | None = None) -> str:
    """Generate a markdown run summary from recipe config + pipeline stats.

    Args:
        recipe: The parsed recipe dict
        stats: Pipeline stats dict (total_source, matched_count, unmatched_count)
        matched_df: The matched DataFrame (for per-step counts)
        timing: Optional pipeline timing dict (load, setup, match, resolve)
        mermaid: Mermaid diagram mode. "default", "detailed", or "disabled"
        recipe_file: Optional recipe filename for header metadata

    Returns:
        Markdown string
    """
    name = recipe.get("name", "Unnamed Recipe")
    desc = recipe.get("description", "")
    total = stats.get("total_source", 0)
    matched = stats.get("matched_count", 0)
    unmatched = stats.get("unmatched_count", 0)
    pct = round(matched / total * 100) if total > 0 else 0

    # Per-step counts: use phase_stats step_counts for multi-phase,
    # fall back to final matched_df grouping for single-phase
    step_counts = {}
    phase_stats_list = stats.get("phases", []) if stats else []
    if phase_stats_list:
        for ps in phase_stats_list:
            step_counts.update(ps.get("step_counts", {}))
    elif matched_df is not None and "match_step" in matched_df.columns:
        for row in matched_df.group_by("match_step").len().iter_rows():
            step_counts[row[0]] = row[1]

    lines = []
    lines.append(f"# {name} -- Run Summary")
    if recipe_file:
        lines.append(f"**Recipe file:** `{recipe_file}`  ")
    if desc:
        lines.append(f"*{desc}*")
    lines.append("")

    # --- Populations ---
    step_sources = {step["source"] for step in _get_all_steps(recipe)}
    step_dests = {step["destination"] for step in _get_all_steps(recipe)}

    lines.append("**Populations:**")
    for pop_name, pop_cfg in _get_all_populations(recipe).items():
        src_name = pop_cfg.get("source", "")

        # Skip _previous_matched populations -- they're runtime-only and
        # showing them with misleading row counts adds confusion
        if src_name == "_previous_matched":
            continue

        source_file = ""
        if src_name and src_name in recipe.get("sources", {}):
            source_file = recipe["sources"][src_name].get("file", "")

        filter_desc = _describe_filters(pop_cfg)
        file_part = f" from {source_file}" if source_file else ""

        if pop_cfg.get("action") == "exclude":
            lines.append(f"- **{pop_name}:** excluded ({filter_desc})")
        elif pop_name in step_sources:
            lines.append(f"- **{pop_name}:** {total} records{file_part} ({filter_desc}) -- *matching target*")
        elif pop_name in step_dests:
            lines.append(f"- **{pop_name}:**{file_part} ({filter_desc}) -- *destination*")
        else:
            lines.append(f"- **{pop_name}:**{file_part} ({filter_desc})")
    lines.append("")

    # --- Results (trailing two spaces for markdown line breaks) ---
    lines.append(f"**Matched:** {matched} of {total} ({pct}%)  ")
    lines.append(f"**Unmatched:** {unmatched} (see Analysis tab)  ")
    if timing:
        lines.append(f"**Timing:** {_format_timing(timing)}  ")
    lines.append("")

    # --- Step table ---
    lines.append("**Matching steps (in priority order):**")
    lines.append("")

    step_header = ("| Step | Source Pop | Source Column | Dest Pop | Dest Column "
                   "| Method | Data Tier | Name Threshold | Address Threshold "
                   "| Address Tier | Date Filter | Other Conditions "
                   "| Matched | % of Total | Leftover |")
    step_sep = "|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|"

    phases = recipe.get("phases", [])
    phase_stats_list = stats.get("phases", []) if stats else []

    if phases:
        # Multi-phase: group steps by phase with phase headers
        step_num = 0
        for phase_idx, phase in enumerate(phases):
            phase_name = phase.get("name", f"Phase {phase_idx + 1}")
            ps = phase_stats_list[phase_idx] if phase_idx < len(phase_stats_list) else {}
            phase_input = ps.get("input_count", "?")
            phase_matched_count = ps.get("matched_count", "?")
            lines.append(f"**{phase_name}** (Input: {phase_input} | Matched: {phase_matched_count})")
            lines.append("")
            lines.append(step_header)
            lines.append(step_sep)

            phase_cumulative = 0
            phase_total = ps.get("input_count", total)
            for step in phase.get("steps", []):
                step_num += 1
                count = step_counts.get(step.get("name", ""), 0)
                phase_cumulative += count
                remaining = phase_total - phase_cumulative
                info = _describe_step_enhanced(
                    step, step_num, count, matched, phase_total, phase_cumulative
                )
                pct = round(count / phase_total * 100, 1) if phase_total > 0 else 0.0
                lines.append(
                    f"| {info['step']} | {info['source_pop']} | {info['source_col']} "
                    f"| {info['dest_pop']} | {info['dest_col']} | {info['method']} "
                    f"| {info['data_tier']} | {info['name_threshold']} "
                    f"| {info['addr_threshold']} "
                    f"| {info['addr_tier']} | {info['date_filter']} "
                    f"| {info['other_conditions']} "
                    f"| {info['matched']} | {pct}% | {remaining} |"
                )
            lines.append("")
    else:
        # Single-phase: flat step list
        lines.append(step_header)
        lines.append(step_sep)

        cumulative = 0
        for i, step in enumerate(_get_all_steps(recipe), 1):
            count = step_counts.get(step.get("name", ""), 0)
            cumulative += count
            remaining = total - cumulative
            info = _describe_step_enhanced(step, i, count, matched, total, cumulative)
            pct_of_total = round(count / total * 100, 1) if total > 0 else 0.0
            lines.append(
                f"| {info['step']} | {info['source_pop']} | {info['source_col']} "
                f"| {info['dest_pop']} | {info['dest_col']} | {info['method']} "
                f"| {info['data_tier']} | {info['name_threshold']} "
                f"| {info['addr_threshold']} "
                f"| {info['addr_tier']} | {info['date_filter']} "
                f"| {info['other_conditions']} "
                f"| {info['matched']} | {pct_of_total}% | {remaining} |"
            )

    lines.append("")
    if recipe.get("phases"):
        lines.append(
            "Each phase passes its matched records to the next phase as input. "
            "Records that match in an early phase but not in a later phase "
            "appear in the final output with null columns for the phases they missed."
        )
    else:
        lines.append(
            "Records that don't match or fail a threshold in one step "
            "move to the next. A record is only unmatched if it fails all steps."
        )

    # --- Mermaid diagram ---
    mermaid_output = None
    if mermaid != "disabled":
        mermaid_output = generate_mermaid(
            recipe, stats, matched_df, detailed=(mermaid == "detailed")
        )
    if mermaid_output:
        lines.append("")
        lines.append("**Matching flow:**")
        lines.append("")
        lines.append("```mermaid")
        lines.append(mermaid_output)
        lines.append("```")

    return "\n".join(lines)


def generate_phase_summary(
    phase_cfg: dict,
    phase_idx: int,
    phase_stats: dict,
    recipe: dict,
    recipe_file: str | None = None,
    mermaid: str = "default",
) -> str:
    """Generate a markdown summary for a single phase in a multi-phase pipeline.

    Produces a self-contained summary with phase-specific stats, step table,
    and Mermaid diagram -- suitable for standalone delivery alongside the
    phase data file.
    """
    phase_name = phase_cfg.get("name", f"Phase {phase_idx + 1}")
    recipe_name = recipe.get("name", "Unnamed Recipe")
    phase_input = phase_stats.get("input_count", 0)
    phase_matched = phase_stats.get("matched_count", 0)
    phase_unmatched = phase_input - phase_matched
    pct = round(phase_matched / phase_input * 100) if phase_input > 0 else 0
    step_counts = phase_stats.get("step_counts", {})

    lines = []
    lines.append(f"# {phase_name} -- Run Summary")
    lines.append(f"**Pipeline:** {recipe_name}")
    if recipe_file:
        lines.append(f"**Recipe file:** `{recipe_file}`  ")
    lines.append("")

    # --- Populations ---
    step_sources = {s["source"] for s in phase_cfg.get("steps", [])}
    step_dests = {s["destination"] for s in phase_cfg.get("steps", [])}

    lines.append("**Populations:**")
    for pop_name, pop_cfg in phase_cfg.get("populations", {}).items():
        src_name = pop_cfg.get("source", "")
        if src_name == "_previous_matched":
            lines.append(f"- **{pop_name}:** output from previous phase")
            continue

        source_file = ""
        if src_name and src_name in recipe.get("sources", {}):
            source_file = recipe["sources"][src_name].get("file", "")

        filter_desc = _describe_filters(pop_cfg)
        file_part = f" from {source_file}" if source_file else ""

        if pop_cfg.get("action") == "exclude":
            lines.append(f"- **{pop_name}:** excluded ({filter_desc})")
        elif pop_name in step_sources:
            lines.append(
                f"- **{pop_name}:** {phase_input} records{file_part} "
                f"({filter_desc}) -- *matching target*"
            )
        elif pop_name in step_dests:
            lines.append(f"- **{pop_name}:**{file_part} ({filter_desc}) -- *destination*")
        else:
            lines.append(f"- **{pop_name}:**{file_part} ({filter_desc})")
    lines.append("")

    # --- Results ---
    lines.append(f"**Matched:** {phase_matched} of {phase_input} ({pct}%)  ")
    lines.append(f"**Unmatched:** {phase_unmatched}  ")
    phase_time = phase_stats.get("time")
    if phase_time is not None:
        lines.append(f"**Phase time:** {phase_time:.2f}s  ")
    lines.append("")

    # --- Step table ---
    lines.append("**Matching steps (in priority order):**")
    lines.append("")

    step_header = (
        "| Step | Source Pop | Source Column | Dest Pop | Dest Column "
        "| Method | Data Tier | Name Threshold | Address Threshold "
        "| Address Tier | Date Filter | Other Conditions "
        "| Matched | % of Total | Leftover |"
    )
    step_sep = "|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|"
    lines.append(step_header)
    lines.append(step_sep)

    cumulative = 0
    for i, step in enumerate(phase_cfg.get("steps", []), 1):
        count = step_counts.get(step.get("name", ""), 0)
        cumulative += count
        remaining = phase_input - cumulative
        info = _describe_step_enhanced(
            step, i, count, phase_matched, phase_input, cumulative
        )
        pct_of_total = round(count / phase_input * 100, 1) if phase_input > 0 else 0.0
        lines.append(
            f"| {info['step']} | {info['source_pop']} | {info['source_col']} "
            f"| {info['dest_pop']} | {info['dest_col']} | {info['method']} "
            f"| {info['data_tier']} | {info['name_threshold']} "
            f"| {info['addr_threshold']} "
            f"| {info['addr_tier']} | {info['date_filter']} "
            f"| {info['other_conditions']} "
            f"| {info['matched']} | {pct_of_total}% | {remaining} |"
        )

    lines.append("")
    lines.append(
        "Records that don't match or fail a threshold in one step "
        "move to the next. A record is only unmatched if it fails all steps."
    )

    # --- Mermaid diagram ---
    if mermaid != "disabled":
        # Build a single-phase Mermaid for this phase
        mini_recipe = {
            "sources": recipe.get("sources", {}),
            "populations": phase_cfg.get("populations", {}),
            "steps": phase_cfg.get("steps", []),
        }
        phase_steps = phase_cfg.get("steps", [])
        mermaid_output = _generate_mermaid_single_phase(
            mini_recipe, phase_steps, step_counts,
            phase_input, phase_matched, phase_unmatched,
            detailed=(mermaid == "detailed"),
        )
        if mermaid_output:
            lines.append("")
            lines.append("**Matching flow:**")
            lines.append("")
            lines.append("```mermaid")
            lines.append(mermaid_output)
            lines.append("```")

    return "\n".join(lines)


def generate_mermaid(recipe: dict, stats: dict, matched_df: pl.DataFrame | None = None,
                     detailed: bool = False) -> str:
    """Generate a Mermaid flowchart from recipe config + pipeline stats.

    Args:
        recipe: The parsed recipe dict
        stats: Pipeline stats dict (total_source, matched_count, unmatched_count)
        matched_df: Optional matched DataFrame (for per-step counts)
        detailed: If True, include thresholds in step boxes (Option E)

    Returns:
        Mermaid flowchart string (without the ```mermaid fences)
    """
    total = stats.get("total_source", 0)
    matched_total = stats.get("matched_count", 0)
    unmatched_total = stats.get("unmatched_count", 0)

    # Per-step counts: prefer phase_stats for multi-phase
    step_counts = {}
    phase_stats_list = stats.get("phases", []) if stats else []
    if phase_stats_list:
        for ps in phase_stats_list:
            step_counts.update(ps.get("step_counts", {}))
    elif matched_df is not None and "match_step" in matched_df.columns:
        for row in matched_df.group_by("match_step").len().iter_rows():
            step_counts[row[0]] = row[1]

    phases = recipe.get("phases", [])
    phase_stats_list = stats.get("phases", []) if stats else []

    if phases and phase_stats_list:
        return _generate_mermaid_multi_phase(
            recipe, phases, phase_stats_list, step_counts,
            total, matched_total, unmatched_total,
        )

    steps = _get_all_steps(recipe)
    if not steps:
        return ""

    return _generate_mermaid_single_phase(
        recipe, steps, step_counts, total, matched_total, unmatched_total,
        detailed=detailed,
    )


def _generate_mermaid_single_phase(
    recipe, steps, step_counts, total, matched_total, unmatched_total,
    detailed=False,
):
    """Mermaid diagram for single-phase cascade recipes."""
    lines = ["flowchart TD"]

    # Source population
    source_pop = steps[0].get("source", "source")
    lines.append(f"    Pop[{source_pop}: {total} records]")

    for i, step in enumerate(steps):
        step_id = f"S{i+1}"
        dest = step.get("destination", "?")
        mf = step.get("match_fields", [{}])[0]
        method = mf.get("method", "?").capitalize()

        if detailed:
            mf.get("threshold", 100 if mf.get("method") == "exact" else "?")
            addr = step.get("address_support", {})
            addr_thresh = addr.get("threshold", "none")
            addr_part = f"addr >= {addr_thresh}%" if isinstance(addr_thresh, (int, float)) else ""

            date_desc = ""
            dg = step.get("date_gate")
            if dg:
                date_desc = f"{dg['field']} < {dg['max_age_years']}yr"
            else:
                for f in step.get("filters", []):
                    if f.get("op") == "max_age_years":
                        date_desc = f"{f['field']} < {f['value']}yr"
                        break

            detail_parts = [p for p in [addr_part, date_desc] if p]
            detail_line = " | ".join(detail_parts)
            if detail_line:
                label = f"{step_id}[\"Step {i+1}: {method} to {dest}\n{detail_line}\"]"
            else:
                label = f"{step_id}[Step {i+1}: {method} to {dest}]"
        else:
            label = f"{step_id}[Step {i+1}: {method} to {dest}]"

        lines.append(f"    {label}")

    lines.append(f"    M[Matched: {matched_total}]")
    lines.append(f"    U[Unmatched: {unmatched_total}]")
    lines.append("")

    remaining = total
    for i, step in enumerate(steps):
        step_id = f"S{i+1}"
        step_name = step.get("name", f"Step {i+1}")
        count = step_counts.get(step_name, 0)

        if i == 0:
            lines.append("    Pop --> S1")

        if count > 0:
            lines.append(f"    {step_id} -->|{count} matched| M")

        remaining -= count

        if i == len(steps) - 1:
            if remaining > 0:
                lines.append(f"    {step_id} -->|{remaining} unmatched| U")
        else:
            lines.append(f"    {step_id} -.->|{remaining} remaining| S{i+2}")

    return "\n".join(lines)


def _generate_mermaid_multi_phase(recipe, phases, phase_stats_list,
                                  step_counts, total, matched_total,
                                  unmatched_total):
    """Mermaid diagram for multi-phase pipelines.

    Shows each phase as a box with input/output counts and arrows
    between phases showing how matched records flow forward.
    """
    lines = ["flowchart TD"]

    # Source node
    source_pop = phases[0].get("steps", [{}])[0].get("source", "source")
    lines.append(f"    Pop[{source_pop}: {total} records]")

    # Phase nodes
    for pi, phase in enumerate(phases):
        ps = phase_stats_list[pi] if pi < len(phase_stats_list) else {}
        phase_name = phase.get("name", f"Phase {pi + 1}")
        p_input = ps.get("input_count", "?")
        p_matched = ps.get("matched_count", "?")

        # Show step breakdown inside the phase
        phase_steps = phase.get("steps", [])
        step_details = []
        for step in phase_steps:
            sname = step.get("name", "?")
            scount = step_counts.get(sname, 0)
            mf = step.get("match_fields", [{}])[0]
            method = mf.get("method", "?").capitalize()
            step_details.append(f"{method}: {scount}")

        detail_str = ", ".join(step_details) if step_details else ""
        if detail_str:
            label = f"P{pi+1}[\"{phase_name}\nIn: {p_input} | Out: {p_matched}\n{detail_str}\"]"
        else:
            label = f"P{pi+1}[\"{phase_name}\nIn: {p_input} | Out: {p_matched}\"]"
        lines.append(f"    {label}")

    lines.append(f"    M[Matched: {matched_total}]")
    lines.append(f"    U[Unmatched: {unmatched_total}]")
    lines.append("")

    # Connections
    lines.append("    Pop --> P1")

    # Unmatched branch from Phase 1 (records that never matched)
    if unmatched_total > 0:
        lines.append(f"    P1 -.->|{unmatched_total} unmatched| U")

    for pi in range(len(phases)):
        ps = phase_stats_list[pi] if pi < len(phase_stats_list) else {}
        p_matched = ps.get("matched_count", 0)
        if pi < len(phases) - 1:
            lines.append(f"    P{pi+1} -->|{p_matched} matched| P{pi+2}")
        else:
            # Last phase feeds into final output
            lines.append(f"    P{pi+1} -->|{matched_total} matched| M")

    return "\n".join(lines)


def write_summary_tab(ws, recipe: dict, stats: dict, matched_df: pl.DataFrame,
                      timing: dict | None = None,
                      recipe_file: str | None = None) -> None:
    """Write a Summary tab to an openpyxl worksheet.

    Args:
        ws: openpyxl Worksheet (already created)
        recipe: The parsed recipe dict
        stats: Pipeline stats dict
        matched_df: The matched DataFrame (for per-step counts)
        timing: Optional pipeline timing dict
        recipe_file: Optional recipe filename for header metadata
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    bold = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    name = recipe.get("name", "Unnamed Recipe")
    desc = recipe.get("description", "")
    total = stats.get("total_source", 0)
    matched = stats.get("matched_count", 0)
    unmatched = stats.get("unmatched_count", 0)
    pct = round(matched / total * 100) if total > 0 else 0

    # Per-step counts: prefer phase_stats for multi-phase
    step_counts = {}
    phase_stats_list_tab = stats.get("phases", []) if stats else []
    if phase_stats_list_tab:
        for ps in phase_stats_list_tab:
            step_counts.update(ps.get("step_counts", {}))
    elif matched_df is not None and "match_step" in matched_df.columns:
        for row in matched_df.group_by("match_step").len().iter_rows():
            step_counts[row[0]] = row[1]

    # Determine population roles
    step_sources = {step["source"] for step in _get_all_steps(recipe)}
    step_dests = {step["destination"] for step in _get_all_steps(recipe)}

    row = 1

    # Title
    ws.cell(row=row, column=1, value=f"{name} -- Run Summary").font = bold
    row += 1
    if recipe_file:
        ws.cell(row=row, column=1, value="Recipe file:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=recipe_file)
        row += 1
    if desc:
        ws.cell(row=row, column=1, value=desc)
        row += 1
    row += 1

    # Populations
    ws.cell(row=row, column=1, value="Populations:").font = Font(bold=True)
    row += 1
    for pop_name, pop_cfg in _get_all_populations(recipe).items():
        src_name = pop_cfg.get("source", "")

        # Skip _previous_matched populations -- runtime-only
        if src_name == "_previous_matched":
            continue

        source_file = ""
        if src_name and src_name in recipe.get("sources", {}):
            source_file = recipe["sources"][src_name].get("file", "")

        filter_desc = _describe_filters(pop_cfg)
        file_part = f" from {source_file}" if source_file else ""

        if pop_cfg.get("action") == "exclude":
            label = f"{pop_name}: excluded ({filter_desc})"
        elif pop_name in step_sources:
            label = f"{pop_name}: {total} records{file_part} ({filter_desc}) -- matching target"
        elif pop_name in step_dests:
            label = f"{pop_name}:{file_part} ({filter_desc}) -- destination"
        else:
            label = f"{pop_name}:{file_part} ({filter_desc})"
        ws.cell(row=row, column=1, value=label)
        row += 1
    row += 1

    # Results
    result_rows = [
        ("Matched", f"{matched} of {total} ({pct}%)"),
        ("Unmatched", f"{unmatched} (see Analysis tab)"),
    ]
    if timing:
        result_rows.append(("Timing", _format_timing(timing)))

    for label, value in result_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Step table
    ws.cell(row=row, column=1, value="Matching steps (in priority order):").font = Font(bold=True)
    row += 1

    headers = [
        "Step", "Source Pop", "Source Column", "Dest Pop", "Dest Column",
        "Method", "Data Tier", "Name Threshold", "Address Threshold",
        "Address Tier", "Date Filter", "Other Conditions",
        "Matched", "% of Total", "Leftover",
    ]

    phases = recipe.get("phases", [])
    phase_stats_list = stats.get("phases", []) if stats else []

    if phases:
        # Multi-phase: group steps by phase
        step_num = 0
        for phase_idx, phase in enumerate(phases):
            phase_name = phase.get("name", f"Phase {phase_idx + 1}")
            ps = phase_stats_list[phase_idx] if phase_idx < len(phase_stats_list) else {}
            phase_input = ps.get("input_count", "?")
            phase_matched_count = ps.get("matched_count", "?")

            # Phase header row
            phase_cell = ws.cell(
                row=row, column=1,
                value=f"{phase_name} (Input: {phase_input} | Matched: {phase_matched_count})"
            )
            phase_cell.font = Font(bold=True)
            row += 1

            # Column headers
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
            row += 1

            phase_cumulative = 0
            phase_total = ps.get("input_count", total)
            for step in phase.get("steps", []):
                step_num += 1
                count = step_counts.get(step.get("name", ""), 0)
                phase_cumulative += count
                remaining = phase_total - phase_cumulative
                info = _describe_step_enhanced(
                    step, step_num, count, matched, phase_total, phase_cumulative
                )
                pct = f"{round(count / phase_total * 100, 1)}%" if phase_total > 0 else "0.0%"
                values = [
                    info["step"], info["source_pop"], info["source_col"],
                    info["dest_pop"], info["dest_col"], info["method"],
                    info["data_tier"], info["name_threshold"], info["addr_threshold"],
                    info["addr_tier"], info["date_filter"], info["other_conditions"],
                    info["matched"], pct, remaining,
                ]
                for ci, v in enumerate(values, 1):
                    ws.cell(row=row, column=ci, value=v)
                row += 1
            row += 1
    else:
        # Single-phase: flat step list
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
        row += 1

        cumulative = 0
        for i, step in enumerate(_get_all_steps(recipe), 1):
            count = step_counts.get(step.get("name", ""), 0)
            cumulative += count
            remaining = total - cumulative
            info = _describe_step_enhanced(step, i, count, matched, total, cumulative)
            pct_of_total = f"{round(count / total * 100, 1)}%" if total > 0 else "0.0%"
            values = [
                info["step"], info["source_pop"], info["source_col"],
                info["dest_pop"], info["dest_col"], info["method"],
                info["data_tier"], info["name_threshold"], info["addr_threshold"],
                info["addr_tier"], info["date_filter"], info["other_conditions"],
                info["matched"], pct_of_total, remaining,
            ]
            for ci, v in enumerate(values, 1):
                ws.cell(row=row, column=ci, value=v)
            row += 1

    row += 1
    if recipe.get("phases"):
        note_text = (
            "Each phase passes its matched records to the next phase as input. "
            "Records that match in an early phase but not in a later phase "
            "appear in the final output with null columns for the phases they missed."
        )
    else:
        note_text = (
            "Records that don't match or fail a threshold in one step "
            "move to the next. A record is only unmatched if it fails all steps."
        )
    note = ws.cell(row=row, column=1, value=note_text)
    note.alignment = wrap
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))

    # Column widths -- derive from header length + padding
    for ci, h in enumerate(headers, 1):
        col_letter = chr(64 + ci) if ci <= 26 else chr(64 + (ci - 1) // 26) + chr(65 + (ci - 1) % 26)
        ws.column_dimensions[col_letter].width = max(len(h) + 4, 10)

