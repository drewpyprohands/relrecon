"""Tests for output.groups tag-only mode (Issue #97)."""

import importlib.util
import json
import sys
from pathlib import Path

import polars as pl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from matching import run_pipeline
from recipe import (
    known_derived_columns,
    load_groups,
    load_recipe,
    validate_fields,
    validate_recipe,
)
from report import apply_output_computations

SRC = Path(__file__).parent.parent / "src"
DATA_DIR = Path(__file__).parent.parent / "data"
RECIPES = Path(__file__).parent / "recipes"
EXPECTED = Path(__file__).parent / "expected"

RECIPE = str(RECIPES / "groups_test.yaml")
FIRST_MATCH_RECIPE = str(RECIPES / "groups_first_match_test.yaml")
UNLISTED_RECIPE = str(RECIPES / "groups_unlisted_test.yaml")
DR_RECIPE = str(RECIPES / "decision_record_test.yaml")


def _load_main():
    spec = importlib.util.spec_from_file_location("relrecon_main", SRC / "__main__.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _run(recipe: dict, tmp_path, out_name="data.csv"):
    """Run the pipeline + _write_output for a recipe dict; return the out dir."""
    result = run_pipeline(recipe, base_dir=str(DATA_DIR))
    main = _load_main()
    main._write_output(
        output_cfg=recipe["output"],
        matched_df=result["matched"],
        unmatched_df=result.get("unmatched"),
        output_path=str(tmp_path / out_name),
        stats=result.get("stats", {}),
        recipe=recipe,
        recipe_file="groups_test.yaml",
        timing=result.get("timing"),
        source_df=result["populations"].get("pop1"),
        source_key="vnd_id",
        base_dir=str(DATA_DIR),
    )
    return tmp_path


def _groups_file(tmp_path, groups, name="groups.json"):
    """Write a groups.json and return its absolute path (literal resolution)."""
    path = tmp_path / name
    path.write_text(json.dumps({"groups": groups}))
    return str(path)


def _cfg(file_path, mode=None):
    cfg = {"file": file_path}
    if mode is not None:
        cfg["mode"] = mode
    return {"groups": cfg}


VALID_GROUP = {
    "group_name": "hyatt",
    "regex": "(?i)hyatt",
    "match_columns": ["l3_fmly_nm"],
}


# ---------------------------------------------------------------------------
# Byte-exact artifacts
# ---------------------------------------------------------------------------

def test_matched_csv_byte_exact(tmp_path):
    """Hyatt regex, values exact-match, and the multi-group row in one file."""
    _run(load_recipe(RECIPE), tmp_path)
    assert (tmp_path / "data.csv").read_text() == (
        EXPECTED / "groups_matched.csv"
    ).read_text()


def test_merged_csv_byte_exact(tmp_path):
    """Unmatched rows carry the same tags as matched ones, from their own values."""
    _run(load_recipe(RECIPE), tmp_path)
    assert (tmp_path / "data_merged.csv").read_text() == (
        EXPECTED / "groups_merged.csv"
    ).read_text()


def test_first_match_matched_csv_byte_exact(tmp_path):
    _run(load_recipe(FIRST_MATCH_RECIPE), tmp_path)
    assert (tmp_path / "data.csv").read_text() == (
        EXPECTED / "groups_first_match_matched.csv"
    ).read_text()


def test_first_match_merged_csv_byte_exact(tmp_path):
    _run(load_recipe(FIRST_MATCH_RECIPE), tmp_path)
    assert (tmp_path / "data_merged.csv").read_text() == (
        EXPECTED / "groups_first_match_merged.csv"
    ).read_text()


# ---------------------------------------------------------------------------
# Matching semantics
# ---------------------------------------------------------------------------

def _merged_tags(tmp_path, recipe_path):
    _run(load_recipe(recipe_path), tmp_path)
    df = pl.read_csv(tmp_path / "data_merged.csv")
    return dict(zip(df["vnd_id"], df["group"], strict=True))


def test_regex_tags_every_hit_across_match_columns(tmp_path):
    """vnd9238 hits on l3 only, vnd3983 on l1 only -- any match_column suffices."""
    tags = _merged_tags(tmp_path, RECIPE)
    assert tags["vnd1234"] == "hyatt"
    assert tags["vnd3983"] == "hyatt"
    assert tags["vnd9238"] == "hyatt"


def test_exclude_regex_kills_the_near_miss(tmp_path):
    """"whyattelsburg" contains "hyatt"; exclude_regex is what untags it."""
    assert _merged_tags(tmp_path, RECIPE)["vnd2364"] == ""


def test_values_match_is_trimmed_and_case_insensitive(tmp_path):
    """" law eagles" hits "Law Eagles"; "Law Eagles Holdings" is not an exact value."""
    tags = _merged_tags(tmp_path, RECIPE)
    assert "law eagles" in tags["vnd2983"]
    assert "law eagles" in tags["vnd3923"]
    assert "law eagles" in tags["vnd7781"]
    assert tags["vnd8420"] == "key accounts"


def test_untagged_row_is_empty(tmp_path):
    assert _merged_tags(tmp_path, RECIPE)["vnd5510"] == ""


def test_multi_group_renders_sorted_and_joined(tmp_path):
    """Sorted alphabetically, not in file order -- law eagles precedes it there."""
    assert _merged_tags(tmp_path, RECIPE)["vnd2983"] == "key accounts;law eagles"


def test_first_match_takes_file_order_not_alphabetical(tmp_path):
    """The discriminating case: file order says law eagles, sorting says key accounts."""
    assert _merged_tags(tmp_path, FIRST_MATCH_RECIPE)["vnd2983"] == "law eagles"


def test_parquet_holds_a_native_list(tmp_path):
    recipe = load_recipe(RECIPE)
    recipe["output"]["format"] = "parquet"
    _run(recipe, tmp_path, out_name="data.parquet")
    df = pl.read_parquet(tmp_path / "data_merged.parquet")
    assert df.schema["group"] == pl.List(pl.String)
    tags = dict(zip(df["vnd_id"], df["group"].to_list(), strict=True))
    assert tags["vnd2983"] == ["key accounts", "law eagles"]
    assert tags["vnd5510"] == []


def test_first_match_parquet_is_a_plain_string(tmp_path):
    recipe = load_recipe(FIRST_MATCH_RECIPE)
    recipe["output"]["format"] = "parquet"
    _run(recipe, tmp_path, out_name="data.parquet")
    df = pl.read_parquet(tmp_path / "data_merged.parquet")
    assert df.schema["group"] == pl.String
    tags = dict(zip(df["vnd_id"], df["group"].to_list(), strict=True))
    assert tags["vnd2983"] == "law eagles"
    assert tags["vnd5510"] is None


def test_xlsx_report_renders_the_joined_tag(tmp_path):
    """openpyxl cannot hold a list cell; the report must see the joined text."""
    from openpyxl import load_workbook

    recipe = load_recipe(RECIPE)
    recipe["output"]["format"] = "xlsx"
    recipe["output"]["summary"] = "xlsx"
    _run(recipe, tmp_path, out_name="data.xlsx")
    ws = load_workbook(tmp_path / "data.xlsx")["Matched"]
    header = [c.value for c in ws[1]]
    col = header.index("group") + 1
    values = [ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1)]
    assert "key accounts;law eagles" in values


# ---------------------------------------------------------------------------
# Visibility and registration
# ---------------------------------------------------------------------------

def test_group_registers_in_known_derived():
    assert "group" in known_derived_columns(load_recipe(RECIPE))


def test_group_absent_from_all_artifacts_when_unlisted(tmp_path):
    """Configured but not named in columns.matched -- these never auto-append."""
    _run(load_recipe(UNLISTED_RECIPE), tmp_path)
    for name in ("data.csv", "data_merged.csv", "data_unmatched.csv"):
        assert "group" not in (tmp_path / name).read_text().splitlines()[0]


def test_no_op_without_groups_leaves_no_group_column(tmp_path):
    """A recipe with no output.groups is untouched by this feature."""
    _run(load_recipe(DR_RECIPE), tmp_path)
    assert "group" not in pl.read_csv(tmp_path / "data_merged.csv").columns


def test_no_op_decision_record_artifacts_unchanged(tmp_path):
    """The Issue #93 fixtures still render byte-for-byte as committed."""
    _run(load_recipe(DR_RECIPE), tmp_path)
    assert (tmp_path / "data_merged.csv").read_text() == (
        EXPECTED / "decision_record_merged.csv"
    ).read_text()


# ---------------------------------------------------------------------------
# Validation: each rule gets a rejecting fixture and a passing control
# ---------------------------------------------------------------------------

def _load_errors(groups, tmp_path, mode=None, raw=None):
    """load_groups error message for a groups list (or raw file text)."""
    if raw is not None:
        path = tmp_path / "groups.json"
        path.write_text(raw)
        cfg = _cfg(str(path), mode)
    else:
        cfg = _cfg(_groups_file(tmp_path, groups), mode)
    with pytest.raises(ValueError) as exc:
        load_groups(cfg, str(DATA_DIR))
    return str(exc.value)


def test_control_valid_groups_file_loads(tmp_path):
    groups, warnings = load_groups(_cfg(_groups_file(tmp_path, [VALID_GROUP])), ".")
    assert [g["group_name"] for g in groups] == ["hyatt"]
    assert warnings == []


def test_reject_missing_file(tmp_path):
    cfg = _cfg(str(tmp_path / "nope.json"))
    with pytest.raises(ValueError, match="not found"):
        load_groups(cfg, str(DATA_DIR))


def test_bare_name_resolves_relative_to_base_dir(tmp_path):
    """Literal first, then base_dir -- the aliases/stopwords convention."""
    _groups_file(tmp_path, [VALID_GROUP], name="side.json")
    groups, _ = load_groups(_cfg("side.json"), str(tmp_path))
    assert groups[0]["group_name"] == "hyatt"


def test_config_prefixed_path_resolves_literally(tmp_path, monkeypatch):
    """`file: config/groups.json` works, exactly like `aliases: config/aliases.json`."""
    cfg_dir = tmp_path / "config"
    cfg_dir.mkdir()
    _groups_file(cfg_dir, [VALID_GROUP], name="groups.json")
    monkeypatch.chdir(tmp_path)
    groups, _ = load_groups(_cfg("config/groups.json"), str(tmp_path / "data"))
    assert groups[0]["group_name"] == "hyatt"


def test_bare_name_with_file_in_config_names_the_fix(tmp_path, monkeypatch):
    """The common mistake: sidecar dropped in config/, referenced bare."""
    cfg_dir = tmp_path / "config"
    cfg_dir.mkdir()
    _groups_file(cfg_dir, [VALID_GROUP], name="groups.json")
    monkeypatch.chdir(tmp_path)
    with pytest.raises(ValueError) as exc:
        load_groups(_cfg("groups.json"), str(tmp_path / "data"))
    msg = str(exc.value)
    assert 'A file exists at "config/groups.json"' in msg
    assert 'write "file: config/groups.json"' in msg


def test_reject_invalid_json(tmp_path):
    assert "not valid JSON" in _load_errors(None, tmp_path, raw="{not json")


def test_reject_empty_groups_list(tmp_path):
    assert 'no non-empty "groups" list' in _load_errors([], tmp_path)


def test_reject_missing_groups_key(tmp_path):
    assert 'no non-empty "groups" list' in _load_errors(
        None, tmp_path, raw=json.dumps({"grups": []})
    )


def test_reject_invalid_regex_names_the_group(tmp_path):
    bad = dict(VALID_GROUP, regex="(?i)(unclosed")
    msg = _load_errors([bad], tmp_path)
    assert "invalid regex" in msg
    assert '"hyatt"' in msg


# ---------------------------------------------------------------------------
# Regex dialect: validation must use the engine that actually matches.
# Python's re and Polars' Rust regex disagree in both directions.
# ---------------------------------------------------------------------------

def test_reject_lookahead_at_load_time(tmp_path):
    """Valid in Python re, rejected by the Rust engine -- must fail at load."""
    bad = dict(VALID_GROUP, regex="(?i)(?=hy)hyatt")
    msg = _load_errors([bad], tmp_path)
    assert "invalid regex" in msg
    assert "look-around" in msg


def test_reject_lookbehind_in_exclude_regex_at_load_time(tmp_path):
    bad = dict(VALID_GROUP, exclude_regex="(?<=w)hyatt")
    msg = _load_errors([bad], tmp_path)
    assert "invalid exclude_regex" in msg
    assert "look-around" in msg


def test_lookahead_never_reaches_the_write_path(tmp_path):
    """The contract is a named load-time error, not a ComputeError mid-write.

    It surfaces as RecipeValidationError -- the same gate every other field
    error goes through -- not as a bare ValueError escaping the loader.
    """
    from recipe import RecipeValidationError

    recipe = load_recipe(RECIPE)
    bad = dict(VALID_GROUP, regex="(?i)(?=hy)hyatt")
    recipe["output"]["groups"]["file"] = _groups_file(tmp_path, [bad])
    out = tmp_path / "out"
    with pytest.raises(RecipeValidationError):
        _run(recipe, out)
    assert not out.exists() or not list(out.glob("*"))


def test_accepts_unicode_class_python_re_rejects(tmp_path):
    r"""\p{Lu} is valid for the Rust engine; validating with re would reject it."""
    good = dict(VALID_GROUP, regex=r"(?i)\p{Lu}+yatt")
    groups, _ = load_groups(_cfg(_groups_file(tmp_path, [good])), ".")
    assert groups[0]["regex"] == r"(?i)\p{Lu}+yatt"


def test_unicode_class_group_tags_rows(tmp_path):
    """And it matches at runtime, proving the accept was not merely permissive."""
    df = pl.DataFrame({"l3_fmly_nm": ["Hyatt", "zzz"]})
    group = dict(VALID_GROUP, regex=r"\p{Lu}yatt")
    assert _tag(df, [group]) == [["hyatt"], []]


def test_reject_invalid_exclude_regex_names_the_group(tmp_path):
    bad = dict(VALID_GROUP, exclude_regex="[z-a]")
    msg = _load_errors([bad], tmp_path)
    assert "invalid exclude_regex" in msg
    assert '"hyatt"' in msg


def test_control_exclude_regex_valid_loads(tmp_path):
    good = dict(VALID_GROUP, exclude_regex="(?i)whyattelsburg")
    groups, _ = load_groups(_cfg(_groups_file(tmp_path, [good])), ".")
    assert groups[0]["exclude_regex"] == "(?i)whyattelsburg"


def test_reject_group_with_neither_regex_nor_values(tmp_path):
    bad = {"group_name": "hyatt", "match_columns": ["l3_fmly_nm"]}
    assert "neither regex nor values" in _load_errors([bad], tmp_path)


def test_control_values_only_group_loads(tmp_path):
    good = {"group_name": "e", "values": ["A"], "match_columns": ["l3_fmly_nm"]}
    groups, _ = load_groups(_cfg(_groups_file(tmp_path, [good])), ".")
    assert groups[0]["values"] == ["A"]


def test_reject_missing_match_columns(tmp_path):
    bad = {"group_name": "hyatt", "regex": "x"}
    assert "no match_columns" in _load_errors([bad], tmp_path)


def test_reject_empty_match_columns(tmp_path):
    bad = dict(VALID_GROUP, match_columns=[])
    assert "no match_columns" in _load_errors([bad], tmp_path)


def test_reject_duplicate_group_name(tmp_path):
    msg = _load_errors([VALID_GROUP, dict(VALID_GROUP)], tmp_path)
    assert "must be unique" in msg
    assert '"hyatt"' in msg


def test_reject_missing_group_name(tmp_path):
    bad = {"regex": "x", "match_columns": ["l3_fmly_nm"]}
    assert "no group_name" in _load_errors([bad], tmp_path)


def test_extra_keys_ignored_with_one_warning_naming_them(tmp_path):
    staged = dict(VALID_GROUP, rollup_parent_id="100049", notes="later")
    groups, warnings = load_groups(_cfg(_groups_file(tmp_path, [staged])), ".")
    assert len(warnings) == 1
    assert "rollup_parent_id" in warnings[0]
    assert "notes" in warnings[0]
    assert groups[0]["group_name"] == "hyatt"


def test_extra_keys_run_succeeds(tmp_path):
    """The committed fixture stages rollup_parent_id; the run still produces output."""
    _run(load_recipe(RECIPE), tmp_path)
    assert (tmp_path / "data.csv").exists()


# ---------------------------------------------------------------------------
# Validation: recipe-level rules
# ---------------------------------------------------------------------------

def _reject(recipe: dict, needle: str):
    with pytest.raises(ValueError) as exc:
        validate_recipe(recipe)
    assert needle in str(exc.value)


def test_reject_invalid_mode():
    """The schema enum fires first and names the path; the Python check in
    validate_recipe is the backstop for when jsonschema is not installed."""
    recipe = load_recipe(RECIPE)
    recipe["output"]["groups"]["mode"] = "best_match"
    _reject(recipe, "$.output.groups.mode")
    _reject(recipe, "is not one of ['all_match', 'first_match']")


def test_control_both_modes_accepted():
    for mode in ("all_match", "first_match"):
        recipe = load_recipe(RECIPE)
        recipe["output"]["groups"]["mode"] = mode
        assert validate_recipe(recipe) is not None


def test_reject_missing_file_key():
    recipe = load_recipe(RECIPE)
    del recipe["output"]["groups"]["file"]
    _reject(recipe, "$.output.groups: 'file' is a required property")


def test_reject_group_in_analysis_columns():
    recipe = load_recipe(RECIPE)
    recipe["output"]["columns"]["analysis"].append(
        {"field": "group", "header": "group"}
    )
    _reject(recipe, "matched-view only")


def test_reject_groups_in_multi_phase():
    recipe = {
        "name": "mp",
        "sources": {"src": {"file": "gr_source.csv"}},
        "phases": [
            {
                "name": "Phase 1",
                "populations": {"pop1": {"source": "src", "record_key": "vnd_id"}},
                "steps": [],
                "output": {
                    "format": "csv",
                    "groups": {"file": "gr_groups.json"},
                },
            }
        ],
    }
    _reject(recipe, "not supported in multi-phase")


def test_reject_derived_column_named_group():
    recipe = load_recipe(RECIPE)
    recipe["steps"][0]["inherit"].append({"source": "dst_l1_id", "as": "group"})
    _reject(recipe, "cannot take a generated column's name")


def test_reject_compare_output_named_group():
    recipe = load_recipe(RECIPE)
    recipe["output"]["compare_columns"] = [
        {"left": "l1_fmly_id", "right": "l1_fmly_id", "output": "group"}
    ]
    _reject(recipe, "collides with the groups output column")


def test_reject_decision_record_write_to_group():
    recipe = load_recipe(RECIPE)
    recipe["output"]["decision_record"] = {
        "candidates": ["l1_fmly_id", "l1_fmly_nm"],
        "write_to": "group",
    }
    _reject(recipe, "reserved for group tagging")


# ---------------------------------------------------------------------------
# Validation: data-aware rules (validate_fields)
# ---------------------------------------------------------------------------

def _field_errors(recipe: dict, extra_source_col: str | None = None):
    from recipe import load_source

    sources = {
        name: load_source(cfg, str(DATA_DIR))
        for name, cfg in recipe["sources"].items()
    }
    if extra_source_col:
        sources["src"] = sources["src"].with_columns(
            pl.lit("x").alias(extra_source_col)
        )
    populations = {
        "pop1": {
            "config": recipe["populations"]["pop1"],
            "df": sources["src"],
            "source": "src",
        }
    }
    return validate_fields(recipe, sources, populations, str(DATA_DIR))


def test_control_fixture_recipe_has_no_field_errors():
    errors, _ = _field_errors(load_recipe(RECIPE))
    assert errors == []


def test_reject_match_columns_absent_from_source(tmp_path):
    recipe = load_recipe(RECIPE)
    bad = dict(VALID_GROUP, match_columns=["nope_nm"])
    recipe["output"]["groups"]["file"] = _groups_file(tmp_path, [bad])
    errors, _ = _field_errors(recipe)
    assert any("nope_nm" in e and "not a source column" in e for e in errors)


def test_reject_group_colliding_with_a_source_column():
    errors, _ = _field_errors(load_recipe(RECIPE), extra_source_col="group")
    assert any("reserved for group tagging" in e for e in errors)


def _groups_field_errors(tmp_path, groups=None, raw=None, ref=None):
    """validate_fields errors for a recipe pointed at a given groups file."""
    recipe = load_recipe(RECIPE)
    if ref is not None:
        recipe["output"]["groups"]["file"] = ref
    elif raw is not None:
        path = tmp_path / "groups.json"
        path.write_text(raw)
        recipe["output"]["groups"]["file"] = str(path)
    else:
        recipe["output"]["groups"]["file"] = _groups_file(tmp_path, groups)
    errors, _ = _field_errors(recipe)
    return errors


def test_bad_groups_file_reports_instead_of_raising(tmp_path):
    """Every groups failure is recipe input, so it belongs in the error report.

    Escaping as a bare ValueError bypasses the [ERROR] summary and --dry-run
    entirely, and reaches the user as a traceback.
    """
    bad = dict(VALID_GROUP, regex="(?i)(?=hy)hyatt")
    errors = _groups_field_errors(tmp_path, [bad])
    assert any("invalid regex" in e for e in errors)


def test_missing_groups_file_reports_instead_of_raising(tmp_path):
    errors = _groups_field_errors(tmp_path, ref=str(tmp_path / "gone.json"))
    assert any("not found" in e for e in errors)


def test_invalid_json_reports_instead_of_raising(tmp_path):
    errors = _groups_field_errors(tmp_path, raw="{not json")
    assert any("not valid JSON" in e for e in errors)


def test_empty_groups_list_reports_instead_of_raising(tmp_path):
    errors = _groups_field_errors(tmp_path, [])
    assert any('no non-empty "groups" list' in e for e in errors)


def test_every_groups_error_surfaces_in_one_pass(tmp_path):
    """One run lists all of them -- not one, fix, rerun, repeat."""
    groups = [
        {"group_name": "a", "regex": "(?=x)a", "match_columns": ["l3_fmly_nm"]},
        {"group_name": "a", "regex": "(?<=y)b", "match_columns": ["nope_col"]},
        {"group_name": "c", "match_columns": ["l3_fmly_nm"]},
    ]
    errors = _groups_field_errors(tmp_path, groups)
    assert any("invalid regex" in e and '"(?=x)a"' in e for e in errors)
    assert any("must be unique" in e for e in errors)
    assert any("neither regex nor values" in e for e in errors)
    assert any("nope_col" in e for e in errors)


def test_extra_keys_warning_surfaces_through_validate_fields():
    _, warnings = _field_errors(load_recipe(RECIPE))
    assert any("rollup_parent_id" in w for w in warnings)


# ---------------------------------------------------------------------------
# Unit-level expression behaviour
# ---------------------------------------------------------------------------

def _tag(df, groups, mode="all_match"):
    cfg = {
        "groups": {"file": "x.json", "mode": mode},
        "columns": {"matched": [{"field": "group", "header": "group"}]},
    }
    return apply_output_computations(df, cfg, groups)["group"].to_list()


def test_absent_match_column_contributes_nothing():
    """The unmatched frame lacks destination columns; the group must not fail."""
    df = pl.DataFrame({"l3_fmly_nm": ["hyatt"]})
    group = dict(VALID_GROUP, match_columns=["l3_fmly_nm", "gone_nm"])
    assert _tag(df, [group]) == [["hyatt"]]


def test_group_with_no_present_match_column_never_applies():
    df = pl.DataFrame({"l3_fmly_nm": ["hyatt"]})
    group = dict(VALID_GROUP, match_columns=["gone_nm"])
    assert _tag(df, [group]) == [[]]


def test_null_source_value_does_not_tag():
    df = pl.DataFrame({"l3_fmly_nm": [None, "hyatt"]}, schema={"l3_fmly_nm": pl.String})
    assert _tag(df, [VALID_GROUP]) == [[], ["hyatt"]]


def test_matching_runs_on_raw_values_only():
    """No normalization: a punctuated variant only matches if the regex says so."""
    df = pl.DataFrame({"l3_fmly_nm": ["H-Y-A-T-T"]})
    assert _tag(df, [VALID_GROUP]) == [[]]


def test_values_case_folding_matches_the_row_side():
    """Both sides fold through Polars. Python casefold() would map the
    configured "STRASSE" onto "strasse" while the row stays "straße",
    silently missing the tag."""
    df = pl.DataFrame({"l3_fmly_nm": ["Straße"]})
    group = {
        "group_name": "s",
        "values": ["STRASSE"],
        "match_columns": ["l3_fmly_nm"],
    }
    assert _tag(df, [group]) == [[]]

    exact = dict(group, values=["STRASSE".replace("SS", "ß")])
    assert _tag(df, [exact]) == [["s"]]


def test_values_case_insensitivity_still_holds_for_plain_ascii():
    df = pl.DataFrame({"l3_fmly_nm": ["  LAW EAGLES "]})
    group = {
        "group_name": "e",
        "values": ["law eagles"],
        "match_columns": ["l3_fmly_nm"],
    }
    assert _tag(df, [group]) == [["e"]]


def test_non_string_column_is_cast_before_matching():
    df = pl.DataFrame({"l3_fmly_nm": [100049]})
    group = {"group_name": "n", "regex": "049", "match_columns": ["l3_fmly_nm"]}
    assert _tag(df, [group]) == [["n"]]
