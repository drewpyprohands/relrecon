"""Integration tests: run actual recipes against test data.

These tests exercise the full pipeline end-to-end -- source loading,
population filtering, matching, inherit, and report generation.
They use small local CSV datasets so they run in seconds.
"""

import os
import subprocess
import sys

import polars as pl
import yaml

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
from matching import run_pipeline  # noqa: E402, I001


DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
RECIPE_DIR = os.path.join(os.path.dirname(__file__), "..", "config", "recipes")


def _load_recipe(name):
    path = os.path.join(RECIPE_DIR, name)
    with open(path) as f:
        return yaml.safe_load(f)


def _run(recipe_name):
    recipe = _load_recipe(recipe_name)
    return run_pipeline(recipe, base_dir=DATA_DIR)


# --- Single-phase recipes ---


class TestTieBreakerExample:
    def test_runs(self):
        result = _run("tie_breaker_example.yaml")
        assert result["stats"]["matched_count"] == 3
        assert result["stats"]["unmatched_count"] == 0

    def test_inherit_columns(self):
        result = _run("tie_breaker_example.yaml")
        assert "derived_supplier_id" in result["matched"].columns


class TestSamePopExample:
    def test_runs(self):
        result = _run("same_pop_example.yaml")
        assert result["stats"]["matched_count"] == 4
        assert result["stats"]["unmatched_count"] == 4


class TestL1Reconciliation:
    def test_runs(self):
        result = _run("l1_reconciliation.yaml")
        assert result["stats"]["matched_count"] == 31
        assert result["stats"]["unmatched_count"] == 14


class TestL1Recon80:
    def test_runs(self):
        result = _run("l1_recon_80.yaml")
        assert result["stats"]["matched_count"] == 36
        assert result["stats"]["unmatched_count"] == 9


class TestNullFilterExample:
    def test_runs(self):
        result = _run("null_filter_example.yaml")
        assert result["stats"]["matched_count"] == 30
        assert result["stats"]["unmatched_count"] == 13


class TestStepDefaultsExample:
    def test_runs(self):
        result = _run("step_defaults_example.yaml")
        assert result["stats"]["matched_count"] == 38
        assert result["stats"]["unmatched_count"] == 7


# --- Multi-phase recipes ---


def _run_cli(recipe_name):
    """Run a recipe through the CLI entry point (for output file generation)."""
    result = subprocess.run(
        ["python3", "-m", "src", "--recipe", os.path.join(RECIPE_DIR, recipe_name),
         "--data", DATA_DIR],
        capture_output=True, text=True,
        cwd=os.path.join(os.path.dirname(__file__), ".."),
    )
    return result


class TestGleifParentLookupTest:
    """Multi-phase pipeline using local test data."""

    def test_runs(self):
        result = _run("gleif_parent_lookup_test.yaml")
        assert result["stats"]["matched_count"] >= 3
        assert "phases" in result["stats"]
        assert len(result["stats"]["phases"]) == 3

    def test_phase_counts(self):
        result = _run("gleif_parent_lookup_test.yaml")
        phases = result["stats"]["phases"]
        # Phase 1: some vendors match LEI
        assert phases[0]["matched_count"] >= 3
        # Phase 2: matched LEIs find parent relationships
        assert phases[1]["matched_count"] >= 3
        # Phase 3: parent LEIs resolve to names
        assert phases[2]["matched_count"] >= 3

    def test_inherit_columns(self):
        """inherit renames _dst2 columns to clean names."""
        result = _run("gleif_parent_lookup_test.yaml")
        matched = result["matched"]
        assert "parent_legal_name" in matched.columns
        assert "parent_country" in matched.columns
        assert "parent_hq_country" in matched.columns
        # The inherited ones should be renamed away from _dst2
        assert "Entity.LegalName_dst2" not in matched.columns

    def test_parent_resolution(self):
        """Verify parent names are correctly resolved."""
        result = _run("gleif_parent_lookup_test.yaml")
        matched = result["matched"].sort("Vendor ID")
        # NVIDIA GRAPHICS PRIVATE LIMITED -> parent NVIDIA CORPORATION
        nvidia = matched.filter(pl.col("Vendor Name") == "NVIDIA GRAPHICS PRIVATE LIMITED")
        if nvidia.height > 0:
            assert nvidia["parent_legal_name"][0] == "NVIDIA CORPORATION"
            assert nvidia["parent_country"][0] == "US"

    def test_phase_snapshots_returned(self):
        """Engine returns per-phase DataFrames."""
        result = _run("gleif_parent_lookup_test.yaml")
        snaps = result.get("phase_snapshots", [])
        assert len(snaps) == 3
        # Phase 1 has more columns than just vendor fields
        assert snaps[0].height >= 3
        assert "LEI" in snaps[0].columns
        # Phase 2 adds relationship columns
        assert "Relationship.EndNode.NodeID" in snaps[1].columns


class TestGleifPhasedOutput:
    """Per-phase output file generation (ADR-003)."""

    def test_phase1_csv_and_md_summary(self):
        """Phase 1: CSV data + markdown summary."""
        r = _run_cli("gleif_phased_output_example.yaml")
        assert r.returncode == 0, r.stderr
        # Phase 1 has auto-generated path (no hardcoded path in recipe)
        assert "Phase 1 data:" in r.stdout
        assert "Phase 1 summary:" in r.stdout
        # Find the CSV file in output
        csv_line = [l for l in r.stdout.splitlines() if "Phase 1 data:" in l][0]
        csv_path = csv_line.split("Phase 1 data: ")[1].split(" (")[0]
        full_csv = os.path.join(os.path.dirname(__file__), "..", csv_path)
        assert os.path.exists(full_csv), f"CSV not found: {full_csv}"
        df = pl.read_csv(full_csv)
        assert df.height >= 3
        assert "Vendor ID" in df.columns
        assert "Matched LEI" in df.columns
        assert "Match Score" in df.columns

    def test_phase2_raw_xlsx_no_summary(self):
        """Phase 2: raw XLSX data, no summary."""
        r = _run_cli("gleif_phased_output_example.yaml")
        assert r.returncode == 0, r.stderr
        assert "Phase 2 data:" in r.stdout
        # Phase 2 has summary: none -- no summary or report lines
        assert "Phase 2 summary:" not in r.stdout
        assert "Phase 2 report:" not in r.stdout
        # Find the XLSX file
        xlsx_line = [l for l in r.stdout.splitlines() if "Phase 2 data:" in l][0]
        xlsx_path = xlsx_line.split("Phase 2 data: ")[1].split(" (")[0]
        full_xlsx = os.path.join(os.path.dirname(__file__), "..", xlsx_path)
        assert os.path.exists(full_xlsx)
        from openpyxl import load_workbook
        wb = load_workbook(full_xlsx)
        ws = wb["Data"]
        headers = [cell.value for cell in ws[1]]
        assert ws.max_row >= 4  # header + 3 data rows
        assert "Child LEI" in headers
        assert "Parent LEI" in headers

    def test_phase3_csv_with_md_and_xlsx_summary(self):
        """Phase 3: CSV data + md summary + xlsx report."""
        r = _run_cli("gleif_phased_output_example.yaml")
        assert r.returncode == 0, r.stderr
        assert "Phase 3 data:" in r.stdout
        assert "Phase 3 summary:" in r.stdout
        assert "Phase 3 report:" in r.stdout

    def test_no_top_level_report(self):
        """Multi-phase recipes must not produce top-level output."""
        r = _run_cli("gleif_phased_output_example.yaml")
        assert r.returncode == 0, r.stderr
        # No top-level "Report saved:" or "Data saved:" lines
        assert "Report saved:" not in r.stdout
        assert "Data saved:" not in r.stdout
