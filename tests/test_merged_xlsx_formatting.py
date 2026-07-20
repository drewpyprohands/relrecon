"""Merged-mode xlsx keeps score rounding and banding (Issue #95).

Regression: generate_report collapsed the merged frame's columns to
(header, header), so _write_data -- which keys both the 2-decimal rounding
and the SCORE_HIGH/MED/LOW banding off the *field* name -- matched neither
and silently skipped both.
"""

import sys
from pathlib import Path

import polars as pl
import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from report import (
    CMP_HIGHER,
    CMP_LOWER,
    SCORE_HIGH,
    SCORE_LOW,
    SCORE_MED,
    generate_report,
)

# Headers deliberately differ from field names -- that is what triggered the bug.
COLUMNS = [
    {"field": "vnd_id", "header": "Source ID"},
    {"field": "match_step", "header": "Match Step"},
    {"field": "name_score", "header": "Name Score"},
    {"field": "addr_score", "header": "Address Score"},
    {"field": "addr_street_score", "header": "Street Score"},
    {"field": "best_rejected_score", "header": "Best Rejected"},
]
SCORE_HEADERS = ["Name Score", "Address Score", "Street Score", "Best Rejected"]

RECIPE = {
    "name": "merged formatting",
    "output": {
        "format": "xlsx",
        "columns": {
            "matched": COLUMNS,
            "analysis": [{"field": "vnd_id", "header": "Source ID"}],
        },
    },
}


def _matched():
    return pl.DataFrame({
        "vnd_id": ["V001", "V002"],
        "match_step": ["Exact L3", "Exact L3"],
        "name_score": [93.84615384615384, 65.11111111111111],   # green, yellow
        "addr_score": [76.12345678901234, 41.98765432109876],   # yellow, red
        "addr_street_score": [88.87654321098765, 55.55555555555555],
        "best_rejected_score": [12.34567890123456, 99.87654321098765],
    })


def _unmatched():
    return pl.DataFrame({"vnd_id": ["V003"]})


def _matched_tab(tmp_path, merged: bool):
    out = tmp_path / f"report_{merged}.xlsx"
    generate_report(_matched(), _unmatched(), str(out), stats=None,
                    recipe=RECIPE, merged=merged)
    ws = load_workbook(out)["Matched"]
    headers = [c.value for c in ws[1]]
    return ws, headers


def _column(ws, headers, name):
    idx = headers.index(name) + 1
    return [ws.cell(row=r, column=idx) for r in range(2, ws.max_row + 1)]


@pytest.mark.parametrize("merged", [False, True])
@pytest.mark.parametrize("header", SCORE_HEADERS)
def test_score_columns_rounded_to_two_decimals(tmp_path, merged, header):
    """Every score field rounds to 2dp in both modes."""
    ws, headers = _matched_tab(tmp_path, merged)
    values = [c.value for c in _column(ws, headers, header) if c.value is not None]
    assert values, f"no values under {header}"
    for v in values:
        assert v == round(v, 2), f"{header} unrounded in merged={merged}: {v}"


@pytest.mark.parametrize("merged", [False, True])
def test_score_banding_applied(tmp_path, merged):
    """name_score/addr_score keep their conditional fills in both modes."""
    ws, headers = _matched_tab(tmp_path, merged)
    expected = {
        "Name Score": [SCORE_HIGH.start_color.rgb, SCORE_MED.start_color.rgb],
        "Address Score": [SCORE_MED.start_color.rgb, SCORE_LOW.start_color.rgb],
    }
    for header, want in expected.items():
        cells = _column(ws, headers, header)
        got = [c.fill.start_color.rgb for c in cells[:2]]
        assert got == want, f"{header} banding lost in merged={merged}: {got}"


def test_merged_matches_non_merged_on_the_matched_rows(tmp_path):
    """The two modes agree cell-for-cell on the rows they share."""
    ws_plain, hdr_plain = _matched_tab(tmp_path, False)
    ws_merged, hdr_merged = _matched_tab(tmp_path, True)

    assert hdr_merged == hdr_plain + ["is_unmatched"]
    for header in SCORE_HEADERS:
        plain = [c.value for c in _column(ws_plain, hdr_plain, header)]
        merged = [c.value for c in _column(ws_merged, hdr_merged, header)]
        assert merged[:len(plain)] == plain, header


def test_unmatched_row_has_no_banding(tmp_path):
    """The appended unmatched row has null scores, so it stays unfilled."""
    ws, headers = _matched_tab(tmp_path, True)
    last = _column(ws, headers, "Name Score")[-1]
    assert last.value is None
    assert last.fill.start_color.rgb == "00000000"


def test_headers_still_come_from_the_recipe(tmp_path):
    """Keeping field names on the frame must not leak them into the header row."""
    _, headers = _matched_tab(tmp_path, True)
    assert "name_score" not in headers
    assert "Name Score" in headers


# ---------------------------------------------------------------------------
# compare_columns banding
# ---------------------------------------------------------------------------

CMP_RECIPE = {
    "name": "compare banding",
    "output": {
        "format": "xlsx",
        "compare_columns": [
            {"left": "a", "right": "b", "output": "revenue_cmp"},
            {"left": "b", "right": "a", "output": "plain_cmp"},
        ],
        "columns": {
            "matched": [
                {"field": "vnd_id", "header": "Source ID"},
                {"field": "match_step", "header": "Match Step"},
                # header differs from field, and one that matches it
                {"field": "revenue_cmp", "header": "Revenue Higher/Lower"},
                {"field": "plain_cmp", "header": "plain_cmp"},
            ],
            "analysis": [{"field": "vnd_id", "header": "Source ID"}],
        },
    },
}


def _cmp_tab(tmp_path, merged: bool):
    matched = pl.DataFrame({
        "vnd_id": ["V001", "V002", "V003"],
        "match_step": ["Exact L3"] * 3,
        "revenue_cmp": ["higher", "lower", "same"],
        "plain_cmp": ["lower", "higher", None],
    })
    unmatched = pl.DataFrame({"vnd_id": ["V004"]})
    out = tmp_path / f"cmp_{merged}.xlsx"
    generate_report(matched, unmatched, str(out), stats=None,
                    recipe=CMP_RECIPE, merged=merged)
    ws = load_workbook(out)["Matched"]
    return ws, [c.value for c in ws[1]]


@pytest.mark.parametrize("merged", [False, True])
@pytest.mark.parametrize("header", ["Revenue Higher/Lower", "plain_cmp"])
def test_compare_banding(tmp_path, merged, header):
    """higher -> light red, lower -> light green, same/empty -> unfilled.

    Covers a header that differs from its field and one that matches it,
    since the merged and non-merged paths label columns differently.
    """
    ws, headers = _cmp_tab(tmp_path, merged)
    cells = _column(ws, headers, header)
    got = [(c.value, c.fill.start_color.rgb) for c in cells]
    for value, rgb in got:
        if value == "higher":
            assert rgb == CMP_HIGHER.start_color.rgb, f"{header} {value} -> {rgb}"
        elif value == "lower":
            assert rgb == CMP_LOWER.start_color.rgb, f"{header} {value} -> {rgb}"
        else:
            assert rgb == "00000000", f"{header} {value!r} should be unfilled"


def test_compare_banding_direction_is_inverted_vs_scores(tmp_path):
    """higher is red and lower is green -- the opposite of the score banding."""
    ws, headers = _cmp_tab(tmp_path, False)
    cells = _column(ws, headers, "Revenue Higher/Lower")
    by_value = {c.value: c.fill.start_color.rgb for c in cells}
    assert by_value["higher"] == SCORE_LOW.start_color.rgb    # light red
    assert by_value["lower"] == SCORE_HIGH.start_color.rgb    # light green


def test_no_compare_banding_when_unconfigured(tmp_path):
    """A recipe without compare_columns leaves ordinary text cells alone."""
    ws, headers = _matched_tab(tmp_path, True)
    for cell in _column(ws, headers, "Match Step"):
        assert cell.fill.start_color.rgb == "00000000"
