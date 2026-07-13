"""Tests for recipe echo (Recipe tab + md section, issue #64)."""

import re
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from openpyxl import Workbook, load_workbook

from matching import run_pipeline
from recipe import load_recipe
from recipe_echo import (
    parse_serialized_recipe,
    read_recipe_tab,
    recipe_md_section,
    serialize_recipe,
    write_recipe_tab,
)
from report import generate_report, write_raw_data

DATA_DIR = Path(__file__).parent.parent / "data"
RECIPES = Path(__file__).parent.parent / "config" / "recipes"
L1_RECIPE = RECIPES / "l1_reconciliation.yaml"
TIE_RECIPE = RECIPES / "tie_breaker_example.yaml"
PHASED_RECIPE = RECIPES / "gleif_phased_output_example.yaml"
STEP_DEFAULTS_RECIPE = RECIPES / "step_defaults_example.yaml"


def _tmp_xlsx():
    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    f.close()
    return f.name


def test_serialize_roundtrips_l1():
    """Resolved recipe with exclusions + derived columns round-trips exactly."""
    recipe = load_recipe(str(L1_RECIPE))
    restored = parse_serialized_recipe(serialize_recipe(recipe))
    assert restored == recipe
    # Feature coverage: exclusions and derived columns survive.
    pops = recipe["populations"]
    assert any(p.get("action") == "exclude" for p in pops.values())
    derived = [
        inh["as"] for step in recipe["steps"] for inh in step.get("inherit", [])
    ]
    assert "derived_l1_name" in derived


def test_serialize_roundtrips_tie_breaker():
    """Recipe with a tie-breaker round-trips exactly."""
    recipe = load_recipe(str(TIE_RECIPE))
    restored = parse_serialized_recipe(serialize_recipe(recipe))
    assert restored == recipe
    assert recipe["output"]["tie_breaker"]["column"] == "supplier_id"


def test_recipe_tab_written_roundtrips_via_worksheet():
    """A standalone worksheet round-trips through write/read helpers."""
    recipe = load_recipe(str(TIE_RECIPE))
    wb = Workbook()
    write_recipe_tab(wb.active, recipe)
    assert read_recipe_tab(wb.active) == recipe


def test_report_has_recipe_tab_and_reconstructs():
    """generate_report adds a Recipe tab that reconstructs the resolved recipe."""
    recipe = load_recipe(str(L1_RECIPE))
    result = run_pipeline(recipe, str(DATA_DIR))
    out = _tmp_xlsx()
    try:
        generate_report(
            result["matched"], result["unmatched"], out,
            stats=result["stats"], recipe=recipe,
        )
        wb = load_workbook(out)
        assert "Recipe" in wb.sheetnames
        assert read_recipe_tab(wb["Recipe"]) == recipe
        wb.close()
    finally:
        Path(out).unlink(missing_ok=True)


def test_no_recipe_tab_without_recipe():
    """Without a recipe, no Recipe tab is written."""
    recipe = load_recipe(str(L1_RECIPE))
    result = run_pipeline(recipe, str(DATA_DIR))
    out = _tmp_xlsx()
    try:
        generate_report(result["matched"], result["unmatched"], out)
        wb = load_workbook(out)
        assert "Recipe" not in wb.sheetnames
        wb.close()
    finally:
        Path(out).unlink(missing_ok=True)


def test_recipe_never_in_csv():
    """Raw CSV export (a DW import) must not contain the recipe echo."""
    recipe = load_recipe(str(L1_RECIPE))
    result = run_pipeline(recipe, str(DATA_DIR))
    with tempfile.TemporaryDirectory() as tmp:
        csv_path = str(Path(tmp) / "matched.csv")
        write_raw_data(result["matched"], csv_path, "csv")
        text = Path(csv_path).read_text()
    assert "# Resolved recipe" not in text
    assert "tie_breaker" not in text


def test_md_section_roundtrips():
    """The markdown recipe section embeds a fenced block that round-trips."""
    recipe = load_recipe(str(TIE_RECIPE))
    section = recipe_md_section(recipe)
    assert section.startswith("## Recipe (resolved)")
    block = re.search(r"```(?:yaml|json)\n(.*)\n```", section, re.DOTALL)
    assert block, "fenced recipe block not found"
    assert parse_serialized_recipe(block.group(1)) == recipe


def test_step_defaults_echo_has_no_yaml_aliases():
    """step_defaults share nested objects; the echo must inline them, not alias."""
    recipe = load_recipe(str(STEP_DEFAULTS_RECIPE))
    text = serialize_recipe(recipe)
    # PyYAML would emit "&id001"/"*id001" for repeated objects without the fix.
    assert "&id" not in text and "*id" not in text
    assert parse_serialized_recipe(text) == recipe


def test_multi_phase_report_recipe_tab_roundtrips():
    """Phase xlsx Recipe tab echoes the full multi-phase recipe (issue #64)."""
    import glob
    import os
    import subprocess

    repo = Path(__file__).parent.parent
    with tempfile.TemporaryDirectory() as tmp:
        env = {**os.environ, "PYTHONPATH": str(repo)}
        subprocess.run(
            [sys.executable, "-m", "src", "--recipe", str(PHASED_RECIPE),
             "--data", str(DATA_DIR)],
            cwd=tmp, env=env, check=True, capture_output=True, text=True,
        )
        reports = glob.glob(str(Path(tmp) / "output" / "phase_3_*_report.xlsx"))
        assert reports, "phase 3 report not generated"

        wb = load_workbook(reports[0])
        assert "Recipe" in wb.sheetnames
        got = read_recipe_tab(wb["Recipe"])
        wb.close()

    assert "phases" in got, "phase Recipe tab lost the top-level phases key"
    assert got == load_recipe(str(PHASED_RECIPE))
