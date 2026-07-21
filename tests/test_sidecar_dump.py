"""Tests for sidecar file dump tabs in the xlsx report (issue #100)."""

import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import polars as pl
from openpyxl import load_workbook

from recipe_echo import read_recipe_tab
from report import generate_report

REPO = Path(__file__).parent.parent
DATA_DIR = REPO / "data"
CONFIG = REPO / "config"

EXCLUSIONS = DATA_DIR / "gr_exclusions.csv"
STOPWORDS = CONFIG / "stopwords.json"
ALIASES = CONFIG / "aliases.json"
GROUPS = DATA_DIR / "gr_groups.json"

SIDECAR_TITLES = ["Exclusions", "Stopwords", "Aliases", "Groups"]


def _tmp_xlsx():
    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    f.close()
    return f.name


def _matched():
    return pl.DataFrame({"vnd_id": ["vnd1234"], "match_step": ["Exact L3"]})


def _recipe(*, exclusions=None, stopwords=None, aliases=None, groups=None):
    """Minimal recipe carrying only the sidecar keys under test."""
    recipe = {"name": "sidecar dump", "steps": []}
    if exclusions:
        recipe["exclusions"] = {"file": exclusions, "id_column": "vnd_id"}
    norm = {}
    if stopwords:
        norm["stopwords"] = stopwords
    if aliases:
        norm["aliases"] = aliases
    if norm:
        recipe["normalization"] = norm
    if groups:
        recipe["output"] = {"groups": {"file": groups, "mode": "all_match"}}
    return recipe


def _sheet_rows(ws):
    """Column A of a dump tab, blank cells normalized back to empty strings."""
    return [
        "" if ws.cell(row=r, column=1).value is None
        else str(ws.cell(row=r, column=1).value)
        for r in range(1, ws.max_row + 1)
    ]


def _report(recipe, base_dir="."):
    path = _tmp_xlsx()
    generate_report(_matched(), None, path, recipe=recipe, base_dir=base_dir)
    return load_workbook(path)


def test_all_four_sidecars_dumped_in_order():
    """All four configured -> four tabs, correct order, path + verbatim lines."""
    sources = {
        "Exclusions": EXCLUSIONS,
        "Stopwords": STOPWORDS,
        "Aliases": ALIASES,
        "Groups": GROUPS,
    }
    wb = _report(_recipe(
        exclusions=str(EXCLUSIONS), stopwords=str(STOPWORDS),
        aliases=str(ALIASES), groups=str(GROUPS),
    ))

    assert wb.sheetnames[-4:] == SIDECAR_TITLES
    assert wb.sheetnames.index("Recipe") < wb.sheetnames.index("Exclusions")

    for title, src in sources.items():
        rows = _sheet_rows(wb[title])
        assert rows[0] == str(src.resolve())
        assert rows[1:] == src.read_text(encoding="utf-8-sig").splitlines()


def test_no_sidecars_leaves_tab_set_unchanged():
    """None configured -> no dump tabs; tab set and cells match the baseline."""
    recipe = _recipe()
    wb = _report(recipe)

    assert not set(wb.sheetnames) & set(SIDECAR_TITLES)
    # xlsx is never byte-stable, so pin the tab set and its cell contents.
    assert wb.sheetnames == ["Matched", "Analysis", "Recipe"]
    assert list(wb["Matched"].values) == [("Match Source",), ("Exact L3",)]
    assert list(wb["Analysis"].values) == [("All records matched",)]
    assert read_recipe_tab(wb["Recipe"]) == recipe


def test_only_groups_configured_adds_only_groups_tab():
    """Partial config -> only that sidecar's tab is added."""
    wb = _report(_recipe(groups=str(GROUPS)))

    assert "Groups" in wb.sheetnames
    assert not {"Exclusions", "Stopwords", "Aliases"} & set(wb.sheetnames)
    assert _sheet_rows(wb["Groups"])[0] == str(GROUPS.resolve())


def test_sidecar_resolved_relative_to_base_dir():
    """Unqualified path falls back to base_dir, matching the loader convention."""
    wb = _report(_recipe(groups="gr_groups.json"), base_dir=str(DATA_DIR))

    rows = _sheet_rows(wb["Groups"])
    assert rows[0] == str(GROUPS.resolve())
    assert rows[1:] == GROUPS.read_text(encoding="utf-8-sig").splitlines()


def test_unreadable_sidecar_skips_tab_with_warning(capsys):
    """Configured but unreadable -> report still generated, tab absent, one WARN."""
    wb = _report(_recipe(
        groups="no_such_groups.json", aliases=str(ALIASES),
    ))

    assert "Groups" not in wb.sheetnames
    assert "Aliases" in wb.sheetnames  # a bad sidecar does not poison the rest
    warnings = [
        line for line in capsys.readouterr().err.splitlines()
        if line.startswith("[WARN]")
    ]
    assert len(warnings) == 1
    assert "no_such_groups.json" in warnings[0]
