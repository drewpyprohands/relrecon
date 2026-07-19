"""Tests for the merged matched/unmatched output view + format list (Issue #81).

Covers build_merged_frame (report.py), the _write_output orchestration
(format list, matched_unmatched modes), the is_unmatched column, and the
new validation rules (reserved step name, multi-phase rejection,
emit_unmatched deprecation/precedence).
"""

import importlib.util
import sys
from pathlib import Path

import polars as pl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from matching import run_pipeline
from recipe import load_recipe, validate_recipe

SRC = Path(__file__).parent.parent / "src"
DATA_DIR = Path(__file__).parent.parent / "data"
RECIPE = str(Path(__file__).parent / "recipes" / "merged_output_test.yaml")

MERGED_HEADER = "vnd_id,l3_fmly_nm,match_step,derived_l1_id,is_unmatched"


def _load_main():
    spec = importlib.util.spec_from_file_location("relrecon_main", SRC / "__main__.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _run(recipe: dict, tmp_path, out_name="data.csv"):
    """Run the pipeline + _write_output for a recipe dict; return the out dir."""
    result = run_pipeline(recipe, base_dir=str(DATA_DIR))
    main = _load_main()
    out = str(tmp_path / out_name)
    main._write_output(
        output_cfg=recipe["output"],
        matched_df=result["matched"],
        unmatched_df=result.get("unmatched"),
        output_path=out,
        stats=result.get("stats", {}),
        recipe=recipe,
        recipe_file="merged_output_test.yaml",
        timing=result.get("timing"),
    )
    return tmp_path


# ---------------------------------------------------------------------------
# Merged view content
# ---------------------------------------------------------------------------

def test_merged_csv_exact(tmp_path):
    """Merged csv matches the spec exactly, with the matched block first.

    Asserts header + column order exactly, the row multiset exactly, and
    that every is_unmatched=false row precedes every is_unmatched=true row.
    Row order WITHIN each block is pipeline-defined and not asserted.
    """
    _run(load_recipe(RECIPE), tmp_path)
    lines = (tmp_path / "data_merged.csv").read_text().splitlines()

    # Header row and column order, exact
    assert lines[0] == MERGED_HEADER

    rows = lines[1:]
    # Row multiset, exact values
    assert sorted(rows) == sorted([
        "V001,Helios Energy,Exact L3,P3100,false",
        "V002,Alpine,Exact L3,P5100,false",
        "V003,Sheyelles,unmatched,,true",
    ])

    # Block order: all matched rows precede all unmatched rows
    flags = [r.rsplit(",", 1)[1] for r in rows]
    assert flags == sorted(flags, key=lambda f: f == "true"), (
        f"is_unmatched=false rows must all precede true rows, got {flags}"
    )
    assert flags.count("false") == 2 and flags.count("true") == 1


def test_merged_block_order_detects_violation():
    """The block-order assertion actually rejects an interleaved frame."""
    flags = ["false", "true", "false"]
    assert flags != sorted(flags, key=lambda f: f == "true")


def test_emits_separate_and_merged_both_formats(tmp_path):
    """format: [csv, parquet] + matched_unmatched: [merged, separate]."""
    _run(load_recipe(RECIPE), tmp_path)
    for ext in ("csv", "parquet"):
        assert (tmp_path / f"data.{ext}").exists()            # matched (separate)
        assert (tmp_path / f"data_unmatched.{ext}").exists()  # unmatched (separate)
        assert (tmp_path / f"data_merged.{ext}").exists()     # merged


def test_separate_filenames_unchanged(tmp_path):
    """Separate matched/unmatched artifacts keep today's names + content."""
    _run(load_recipe(RECIPE), tmp_path)
    matched = pl.read_csv(tmp_path / "data.csv")
    assert matched.columns == ["vnd_id", "l3_fmly_nm", "match_step", "derived_l1_id"]
    assert matched.height == 2
    unmatched = pl.read_csv(tmp_path / "data_unmatched.csv")
    assert unmatched["vnd_id"].to_list() == ["V003"]


def test_is_unmatched_only_in_merged(tmp_path):
    _run(load_recipe(RECIPE), tmp_path)
    assert "is_unmatched" in pl.read_csv(tmp_path / "data_merged.csv").columns
    assert "is_unmatched" not in pl.read_csv(tmp_path / "data.csv").columns
    assert "is_unmatched" not in pl.read_csv(tmp_path / "data_unmatched.csv").columns


def test_merged_parquet_matches_csv(tmp_path):
    _run(load_recipe(RECIPE), tmp_path)
    pq = pl.read_parquet(tmp_path / "data_merged.parquet")
    assert pq.columns == MERGED_HEADER.split(",")
    assert pq.height == 3
    assert pq.filter(pl.col("is_unmatched"))["match_step"].to_list() == ["unmatched"]


def test_zero_unmatched(tmp_path):
    """V003 removed -> merged still emitted, all is_unmatched false."""
    recipe = load_recipe(RECIPE)
    recipe["sources"]["src"]["file"] = "mu_source_zero.csv"
    _run(recipe, tmp_path)
    merged = pl.read_csv(tmp_path / "data_merged.csv")
    assert merged.height == 2
    assert merged["is_unmatched"].to_list() == [False, False]
    assert merged.columns == MERGED_HEADER.split(",")


def test_all_unmatched(tmp_path):
    """Dest emptied -> merged has all source rows as sentinel/empty."""
    recipe = load_recipe(RECIPE)
    recipe["sources"]["dst"]["file"] = "mu_dest_empty.csv"
    _run(recipe, tmp_path)
    merged = pl.read_csv(tmp_path / "data_merged.csv")
    assert merged.height == 3
    assert merged["is_unmatched"].to_list() == [True, True, True]
    assert merged["match_step"].to_list() == ["unmatched"] * 3
    assert merged["derived_l1_id"].to_list() == [None, None, None]


def test_merged_only_ignores_emit_unmatched(tmp_path):
    """matched_unmatched: [merged] emits only merged (emit_unmatched ignored)."""
    recipe = load_recipe(RECIPE)
    recipe["output"]["matched_unmatched"] = ["merged"]
    recipe["output"]["format"] = "csv"
    recipe["output"]["emit_unmatched"] = True  # deprecated -- must be ignored
    _run(recipe, tmp_path)
    assert (tmp_path / "data_merged.csv").exists()
    assert not (tmp_path / "data.csv").exists()
    assert not (tmp_path / "data_unmatched.csv").exists()


def test_all_unmatched_xlsx(tmp_path):
    """All-unmatched xlsx: matched frame is empty (0,0) -- merged must still
    write every unmatched row into the Matched tab (regression: the width>0
    guard silently emitted zero rows)."""
    from openpyxl import load_workbook

    recipe = load_recipe(RECIPE)
    recipe["sources"]["dst"]["file"] = "mu_dest_empty.csv"
    recipe["output"]["format"] = "xlsx"
    recipe["output"]["summary"] = ["xlsx"]
    recipe["output"]["matched_unmatched"] = "merged"
    _run(recipe, tmp_path, out_name="data.xlsx")

    ws = load_workbook(tmp_path / "data.xlsx")["Matched"]
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    assert rows[0] == MERGED_HEADER.split(",")  # merged columns intact
    data = rows[1:]
    assert len(data) == 3                                    # all rows present
    assert [r[-1] for r in data] == [True, True, True]        # is_unmatched
    assert [r[2] for r in data] == ["unmatched"] * 3          # match_step
    assert [r[3] for r in data] == [None, None, None]         # derived empty
    assert sorted(r[0] for r in data) == ["V001", "V002", "V003"]


def test_xlsx_merged_appends_into_match_tab(tmp_path):
    """xlsx merged: unmatched rows land in the (unrenamed) Matched tab."""
    from openpyxl import load_workbook

    recipe = load_recipe(RECIPE)
    recipe["output"]["format"] = "xlsx"
    recipe["output"]["summary"] = ["xlsx"]
    recipe["output"]["matched_unmatched"] = "merged"
    _run(recipe, tmp_path, out_name="data.xlsx")

    wb = load_workbook(tmp_path / "data.xlsx")
    assert "Matched" in wb.sheetnames  # tab name unchanged
    ws = wb["Matched"]
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    assert rows[0][-1] == "is_unmatched"
    unmatched_rows = [r for r in rows[1:] if r[0] == "V003"]
    assert len(unmatched_rows) == 1
    assert unmatched_rows[0][2] == "unmatched"  # match_step sentinel
    assert unmatched_rows[0][-1] is True         # is_unmatched flag


# ---------------------------------------------------------------------------
# No-op guarantee (legacy path unchanged)
# ---------------------------------------------------------------------------

def test_noop_single_format_no_matched_unmatched(tmp_path):
    """No matched_unmatched + single-string format -> only the matched file."""
    recipe = load_recipe(RECIPE)
    recipe["output"]["format"] = "csv"
    del recipe["output"]["matched_unmatched"]
    recipe["output"]["summary"] = "none"
    _run(recipe, tmp_path)
    assert (tmp_path / "data.csv").exists()
    assert not (tmp_path / "data_merged.csv").exists()
    assert not (tmp_path / "data_unmatched.csv").exists()


def test_output_path_verbatim_no_extension(tmp_path):
    """`--output data` (no extension) writes a file literally named `data`."""
    recipe = load_recipe(RECIPE)
    recipe["output"]["format"] = "csv"
    del recipe["output"]["matched_unmatched"]
    recipe["output"]["summary"] = "none"
    _run(recipe, tmp_path, out_name="data")
    assert (tmp_path / "data").is_file()
    assert not (tmp_path / "data.csv").exists()


def test_output_path_extension_not_rederived(tmp_path):
    """`--output data.csv` with format: parquet writes `data.csv` (main-exact)."""
    recipe = load_recipe(RECIPE)
    recipe["output"]["format"] = "parquet"
    del recipe["output"]["matched_unmatched"]
    recipe["output"]["summary"] = "none"
    _run(recipe, tmp_path, out_name="data.csv")
    assert (tmp_path / "data.csv").is_file()
    assert not (tmp_path / "data.parquet").exists()
    # Content is parquet despite the .csv name -- the format drives the writer
    assert pl.read_parquet(tmp_path / "data.csv").height == 2


def test_format_list_does_not_collide_on_output_path(tmp_path):
    """A list derives a path per format, so no format overwrites another."""
    recipe = load_recipe(RECIPE)
    recipe["output"]["format"] = ["parquet", "csv"]
    del recipe["output"]["matched_unmatched"]
    recipe["output"]["summary"] = "none"
    _run(recipe, tmp_path, out_name="data.csv")
    assert pl.read_parquet(tmp_path / "data.parquet").height == 2
    assert pl.read_csv(tmp_path / "data.csv").height == 2


# ---------------------------------------------------------------------------
# Validation rules
# ---------------------------------------------------------------------------

def test_reserved_step_name_unmatched():
    recipe = load_recipe(RECIPE)
    recipe["steps"][0]["name"] = "unmatched"
    with pytest.raises(ValueError, match="reserved"):
        validate_recipe(recipe)


def test_multi_phase_matched_unmatched_rejected():
    recipe = load_recipe(str(Path(__file__).parent / "recipes" / "gleif_parent_phased_test.yaml"))
    idx = next(i for i, p in enumerate(recipe["phases"]) if "output" in p)
    recipe["phases"][idx]["output"]["matched_unmatched"] = "merged"
    with pytest.raises(ValueError, match="multi-phase"):
        validate_recipe(recipe)


def test_emit_unmatched_deprecation_warning():
    recipe = load_recipe(RECIPE)  # matched_unmatched already present
    recipe["output"]["emit_unmatched"] = True
    warnings = validate_recipe(recipe)
    assert any("emit_unmatched is deprecated and ignored" in w for w in warnings)


def test_emit_unmatched_alone_no_warning():
    recipe = load_recipe(RECIPE)
    del recipe["output"]["matched_unmatched"]
    recipe["output"]["format"] = "csv"
    recipe["output"]["summary"] = "none"
    recipe["output"]["emit_unmatched"] = True
    warnings = validate_recipe(recipe)
    assert not any("emit_unmatched" in w for w in warnings)


def test_is_unmatched_reserved_as_field():
    recipe = load_recipe(RECIPE)
    recipe["output"]["columns"]["matched"][0]["field"] = "is_unmatched"
    with pytest.raises(ValueError, match="is_unmatched.*reserved"):
        validate_recipe(recipe)


def test_is_unmatched_reserved_as_header():
    recipe = load_recipe(RECIPE)
    recipe["output"]["columns"]["matched"][0]["header"] = "is_unmatched"
    with pytest.raises(ValueError, match="is_unmatched.*reserved"):
        validate_recipe(recipe)


def test_is_unmatched_reserved_in_analysis_columns():
    recipe = load_recipe(RECIPE)
    recipe["output"]["columns"]["analysis"][0]["header"] = "is_unmatched"
    with pytest.raises(ValueError, match="is_unmatched.*reserved"):
        validate_recipe(recipe)


def test_is_unmatched_allowed_elsewhere():
    """Negative control: the reserved check is scoped to output.columns."""
    recipe = load_recipe(RECIPE)  # no is_unmatched anywhere
    assert validate_recipe(recipe) is not None  # validates without raising
    # A similarly-named column is not caught by the reserved check
    recipe["output"]["columns"]["matched"][0]["header"] = "is_unmatched_flag"
    validate_recipe(recipe)


def test_schema_accepts_format_list_and_matched_unmatched():
    recipe = load_recipe(RECIPE)  # format: [csv, parquet], matched_unmatched: [merged, separate]
    warnings = validate_recipe(recipe)
    assert not any("additionalProperties" in w for w in warnings)
